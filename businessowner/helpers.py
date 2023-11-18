import calendar
from .models import *
from .schemas import *
from django.http import JsonResponse
from django.shortcuts import redirect
from ninja import File
from django.db.models import Q
from ninja.errors import HttpError
from .utils import generate_token,paginate_data
import pandas as pd
from datetime import timedelta, timezone
import random
import time
import jwt
import razorpay
from datetime import datetime
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils import timezone
import base64
from urllib.request import urlopen
from django.core.files.base import ContentFile  # Import ContentFile
import os
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from django.http import HttpResponse
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.units import inch
from django.db.models import F
from PIL import Image
from io import BytesIO
import openpyxl


from openpyxl_image_loader import SheetImageLoader
import boto3
from dotenv import load_dotenv
import pathlib
s3_client = boto3.client(
        's3',
        aws_access_key_id='AKIAS5UGKMCVJSQA2DU7',
        aws_secret_access_key='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',
        region_name='ap-south-1'
    )

def perform_login(data):
    try:
        user = BusinessOwners.objects.get(email=data.email)
        if user.password == data.password:
            token = generate_token(str(user.id))
            # if timezone.now() > plan_purchased.expire_date:
            #     user.is_plan_purchase = False
            #     user.save()
            if user.is_plan_purchase == False:
                city = Cities.objects.get(id=user.city_id)
                state = States.objects.get(id=city.state_id)
                response_data = {
                    "result": True,
                    "data": {
                        "id": str(user.id),
                        "business_name": user.business_name,
                        "business_type": user.business_type,
                        "first_name": user.first_name,
                        "last_name": user.last_name,
                        "email": user.email,
                        "contact_no": user.contact_no,
                        "address": user.address,
                        "logo": user.logo.url if user.logo else None,
                        "tuition_tagline": user.tuition_tagline if user.tuition_tagline else None,
                        "status": user.status,
                        "is_reset":user.is_reset,
                        "is_plan_purchased":user.is_plan_purchase,
                        "city": {
                            "city_id": str(city.id),
                            "city_name": city.name,
                            "state_id": str(city.state_id),
                            "state_name": state.name,
                        },
                        "token": token
                        },
                    "message": "Login successful",
            
                }
                return response_data
            else:
                plan_purchased = PurchaseHistory.objects.filter(business_owner=user, status__in=[True]).order_by('-start_date')[:1].first()

                if timezone.now() > plan_purchased.expire_date:
                    user.is_plan_purchase = False
                    user.save()

                city = Cities.objects.get(id=user.city_id)
                state = States.objects.get(id=city.state_id)
                response_data = {
                    "result": True,
                    "data": {
                        "id": str(user.id),
                        "business_name": user.business_name,
                        "business_type": user.business_type,
                        "first_name": user.first_name,
                        "last_name": user.last_name,
                        "email": user.email,
                        "contact_no": user.contact_no,
                        "address": user.address,
                        "logo": user.logo.url if user.logo else None,
                        "tuition_tagline": user.tuition_tagline if user.tuition_tagline else None,
                        "status": user.status,
                        "is_reset":user.is_reset,
                        "is_plan_purchased":user.is_plan_purchase,
                        "city": {
                            "city_id": str(city.id),
                            "city_name": city.name,
                            "state_id": str(city.state_id),
                            "state_name": state.name,
                        },
                        "token": token
                        },
                    "message": "Login successful",
            
                }
            return response_data

        else:
            response_data = {
            "result": False,
            "message": "Invalid Password"
        }
    
        return JsonResponse(response_data, status=401)
    except BusinessOwners.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Invalid Email"
        }
    
        return JsonResponse(response_data, status=401)
    
    except Exception as e:
        response_data = {
            "result": False,
            "message": "Something went wrong"
        }
        return JsonResponse(response_data, status=401)



def perform_change_password(data, user):
    try:
        if data.old_password != user.password:
            response_data = {"result": False, "message": "Invalid old password"}
            return JsonResponse(response_data, status=400)
        
        if data.new_password == user.password:
            response_data = {"result": False, "message": "Please change password"}
            return JsonResponse(response_data, status=400)
        
        user.password = data.new_password
        user.is_reset = True
        user.save()

        response_data = {"result": True, "message": "Password changed successfully"}
        return JsonResponse(response_data, status=200)
    
    except BusinessOwners.DoesNotExist:
        response_data = {"result": False, "message": "Business owner not found"}
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
            "result": False,
            "message": "Something went wrong" 
        }
        return JsonResponse(response_data, status=400)


def perform_forgot_password(data):
    try:
        email = data.email
        user = BusinessOwners.objects.get(email=email)
        subject = 'Acount Recovery'
        token = generate_token(str(user.id))
        link = "api/businessOwner/resetPasswordLink/"

        html_message = render_to_string('forgot_password.html', {'token': token, 'link':link})

        from_email = settings.DEFAULT_FROM_EMAIL
        to_email = [email]
        
        email = EmailMultiAlternatives(subject, body=None, from_email=from_email, to=to_email)
        email.attach_alternative(html_message, "text/html")
        email.send()
        response_data = {
            "result": True,
            "message":"Email sent successfully"
        }
        return response_data
    
    except BusinessOwners.DoesNotExist:
        response_data = {"result": False, "message": "Business owner not found"}
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
            "result": False,
            "message": "Something went wrong" 
        }
        return JsonResponse(response_data, status=400)


def verify_reset_password_link(token):
    try:
        payload = jwt.decode(token, 'secret', algorithms=['HS256'])
        return redirect('http://localhost:3000/resetPassword/${token}')
    
    except (jwt.DecodeError, jwt.ExpiredSignatureError):
        response_data = {
            "result": False,
            "message": "Reset Link Expired" 
        }
        return JsonResponse(response_data, status=400)
       
    except Exception as e:
        response_data = {
            "result": False,
            "message": "Something went wrong" 
        }
        return JsonResponse(response_data, status=400)


def perform_reset_password(data):
    try:
        payload = jwt.decode(data.token, 'secret', algorithms=['HS256'])
        user = BusinessOwners.objects.get(id=payload["user_id"])
        if data.new_password == data.confirm_password:
            user.password = data.new_password
            user.save()
        response_data = {
            "result":True,
            "message": "Password rest"
        }
        return response_data

    except BusinessOwners.DoesNotExist:
        response_data = {"result": False, "message": "Business owner not found"}
        return JsonResponse(response_data, status=400)
    
    except Exception as e:
        response_data = {
            "result": False,
            "message": "Something went wrong" 
        }
        return JsonResponse(response_data, status=400)


#-----------------------------------------------------------------------------------------------------------#
#----------------------------------------------CITY & STATE-------------------------------------------------#
#-----------------------------------------------------------------------------------------------------------#


def get_citylist(request):
    try:
        city_list = [] 
        cities = Cities.objects.all()
        for city in cities:
            city_dict = {
                "city_id": str(city.id),
                "city_name": city.name,
                "state_id": str(city.state.id),
                "state_name": city.state.name
            }
            city_list.append(city_dict)  
        
        paginated_city_data, items_per_page = paginate_data(request, city_list)

        return {
            "result": True,
            "data": list(paginated_city_data),
            "message": "Data retrieved successfully",   
            "pagination": {
                "page": paginated_city_data.number,
                "total_docs": paginated_city_data.paginator.count,
                "total_pages": paginated_city_data.paginator.num_pages,
                "per_page": items_per_page,
            },
        }  

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)

def get_statelist(request):
    try:
        state_list =[]
        states = States.objects.all()
        for state in states:
            state_dict = {
                "state_id": str(state.id),
                "state_name": state.name
            }
            state_list.append(state_dict)

        paginated_state_data, items_per_page = paginate_data(request, state_list)

        return {
            "result": True,
            "data": list(paginated_state_data),
            "message": "Data retrieved successfully",   
            "pagination": {
                "page": paginated_state_data.number,
                "total_docs": paginated_state_data.paginator.count,
                "total_pages": paginated_state_data.paginator.num_pages,
                "per_page": items_per_page,
            },
        }
    
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)


#-----------------------------------------------------------------------------------------------------------#
#---------------------------------------------PLAN PURCHASE-------------------------------------------------#
#-----------------------------------------------------------------------------------------------------------#


def get_plan_list(request):
    try:
        plans = Plans.objects.all()

        plan_schema_list = [
            {
            "id":str(plan.id),
            "plan_name":plan.plan_name,
            "description":plan.description,
            "price":plan.price,
            "validity":plan.validity,
            "image":plan.image.url if plan.image else None,
            "status":plan.status
        } for plan in plans]

        paginated_plan_data, items_per_page = paginate_data(request, plan_schema_list)

        return {
            "result": True,
            "data": list(paginated_plan_data),
            "message": "Data retrieved successfully",   
            "pagination": {
                "page": paginated_plan_data.number,
                "total_docs": paginated_plan_data.paginator.count,
                "total_pages": paginated_plan_data.paginator.num_pages,
                "per_page": items_per_page,
            },
        }

    except Plans.DoesNotExist:
        response_data = {"result": False, "message": "Plans not found"}
        return JsonResponse(response_data, status=400)
    except Exception as e:
        response_data = {
            "result": False,
            "message": "Something went wrong"
        }
        return JsonResponse(response_data, status=400)
    

def purchase_plan(data, user):
    try:  
        
        plan = Plans.objects.get(id=data.id)
        business_user = BusinessOwners.objects.get(id=user.id)
        purchases = PurchaseHistory.objects.filter(plan=data.id, business_owner=user)
        for purchase in purchases:
            if purchase.expire_date > timezone.now():
                response_data = {
                "result": False,
                "message": f"Plan has already purchased for {plan.validity} months."
            }
            return JsonResponse(response_data, status=400)

                
        client = razorpay.Client(auth=(settings.RAZORPAY_API_KEY, settings.RAZORPAY_API_SECRET))
        order_data = {
            "amount": plan.price*100,
            "currency":'INR'
        }

        order = client.order.create(order_data)
        order_id = order.get('id', '')
        purchase_history = PurchaseHistory(
            plan=plan,
            business_owner=user,  
            order_id=order_id
        )
        purchase_history.save()
        purchase_data = {
            "order_id": order_id,
            "price": plan.price,
            "validity":plan.validity,
            "description":plan.description,
            "plan":purchase_history.plan.plan_name,
            "first_name":business_user.first_name,
            "last_name":business_user.last_name,
            "contact_no":business_user.contact_no,
            "email":business_user.email


        }
        response_data = {
            "result": True,
            "data":purchase_data,
            "message": "Order-id generated"
        }
        return response_data

       
    except Exception as e:
        response_data = {
            "result": False,
            "message": "Something went wrong"
        }
        return JsonResponse(response_data, status=400)
    

def verify_plan_payment(data, user):
    try:
        purchased_plan = PurchaseHistory.objects.get(order_id=data.id)
        purchased_plan.status = True
        purchased_plan.start_date = datetime.now()
       
        purchased_plan.save()
        duration = purchased_plan.plan.validity
        expire_date = purchased_plan.start_date + timedelta(days=calendar.monthrange(purchased_plan.start_date.year, purchased_plan.start_date.month)[1] * duration)

        purchased_plan.expire_date = expire_date
        purchased_plan.save()

        client = BusinessOwners.objects.get(id=user.id)
        client.is_plan_purchase = True
        client.save()

        plan_data = {
            "order_id": purchased_plan.order_id,
            "status": purchased_plan.status,
            "purched_plan":purchased_plan.plan.plan_name,
            "is_plan_purchased":client.is_plan_purchase
        }

        response_data = {
            "result": True,
            "data": plan_data,
            "message": "Plan purchased successfully"
        }
        return response_data

    except Exception as e:
        response_data = {
            "result": False,
            "message": "Something went wrong"
        }
        return JsonResponse(response_data, status=400)


def get_purchase_history(request):
    try:
        purchase_history = PurchaseHistory.objects.filter(business_owner=request.user, status__in=[True])

        purchase_history_list = [
            {
                "plan_name": purchase.plan.plan_name,
                "order_id": purchase.order_id,
                "status": purchase.status,
                "purchase_date":purchase.start_date
         
            }
            for purchase in purchase_history
        ]
        paginated_plan_purchase, items_per_page = paginate_data(request, purchase_history_list)

        return {
            "result": True,
            "data": list(paginated_plan_purchase),
            "message": "Data retrieved successfully",   
            "pagination": {
                "page": paginated_plan_purchase.number,
                "total_docs": paginated_plan_purchase.paginator.count,
                "total_pages": paginated_plan_purchase.paginator.num_pages,
                "per_page": items_per_page,
            },
        }
    
    except Exception as e:
        response_data = {
            "result": False,
            "message": "Something went wrong"
        }
        return JsonResponse(response_data, status=400)


#-----------------------------------------------------------------------------------------------------------#
#------------------------------------------------DASHBOARD--------------------------------------------------#
#-----------------------------------------------------------------------------------------------------------#        
    

def dashboard(user):
    try:
        b_type = user.business_type
        
        if b_type not in ["competitive", "academic"]:
            return JsonResponse({
                "result": False,
                "message": "Invalid business type"
            }, status=400)

        data = []

        if b_type == "competitive":
            data.append({"name": "Exams", "count": CompetitiveExams.objects.filter(business_owner=user, start_date__isnull=False).count()})
            data.append({"name": "Students", "count": Students.objects.filter(business_owner=user).count()})
            data.append({"name": "Batches", "count": CompetitiveBatches.objects.filter(business_owner=user).count()})
            data.append({"name": "Subjects", "count": CompetitiveSubjects.objects.filter(business_owner=user).count()})
            latest_exams = CompetitiveExams.objects.filter(business_owner=user, start_date__isnull=False).order_by('-start_date')[:5]
        else:
            data.append({"name": "Exams", "count": AcademicExams.objects.filter(business_owner=user, start_date__isnull=False).count()})
            data.append({"name": "Students", "count": Students.objects.filter(business_owner=user).count()})
            data.append({"name": "Boards", "count": AcademicBoards.objects.filter(business_owner=user).count()})
            data.append({"name": "Mediums", "count": AcademicMediums.objects.filter(board_name__business_owner=user).count()})
            data.append({"name": "Standards", "count": AcademicStandards.objects.filter(medium_name__board_name__business_owner=user).count()})
            data.append({"name": "Subjects", "count": AcademicSubjects.objects.filter(standard__medium_name__board_name__business_owner=user).count()})
            latest_exams = AcademicExams.objects.filter(business_owner=user, start_date__isnull=False).order_by('-start_date')[:5]

        latest_exam_data = [
            {
                "id": exam.id,
                "title": exam.exam_title,
                "start_date": exam.start_date.strftime("%Y-%m-%d"),
            }
            for exam in latest_exams
        ]

        data.append({"name": "Latest Exams", "exams": latest_exam_data})

        response_data = {
            "result": True,
            "data": data,
            "message": "Dashboard data retrieved successfully"
        }

        return response_data

    except Exception as e:
        print(e)
        response_data = {
            "result": False,
            "message": "Something went wrong"
        }
        return JsonResponse(response_data, status=400)
    

#-----------------------------------------------------------------------------------------------------------#
#---------------------------------------------OWNER PROFILE-------------------------------------------------#
#-----------------------------------------------------------------------------------------------------------#


def create_owner_response(user, is_valid, message):
    try:
        owner = BusinessOwners.objects.get(id=user.id)

        if owner:
            city = Cities.objects.get(id=owner.city_id)
            state = States.objects.get(id=city.state_id)
        
            owner_data = {
                    "id": str(owner.id),
                    "business_name": owner.business_name,
                    "business_type": owner.business_type,
                    "first_name": owner.first_name,
                    "last_name": owner.last_name,
                    "email": owner.email,
                    "contact_no": owner.contact_no,
                    "address": owner.address,
                    "logo": owner.logo.url if owner.logo else None,
                    "tuition_tagline": owner.tuition_tagline if owner.tuition_tagline else None,
                    "status": owner.status,
                    "created_at": owner.created_at,
                    "city": {
                        "city_id": str(city.id),
                        "city_name": city.name,
                        "state_id": str(city.state_id),
                        "state_name": state.name,
                    },
                }

            response_data = {
                "result": is_valid,
                "data": owner_data,
                "message": message
            }
        else:
            response_data = {
                "result": False,
                "message": "Owner not found"
            }
        
        return response_data
    

    except BusinessOwners.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Business owner does not exist"
        }
        return JsonResponse(response_data, status=400)


def update_owner_data(data, user):
    try:
        owner = BusinessOwners.objects.get(id=user.id)
        if owner:
            update_data = {field: value for field, value in data.dict().items() if value is not None}
            if update_data:
                for field, value in update_data.items():
                    if field == "city":
                        city = Cities.objects.get(id=value)
                        owner.city = city
                    elif field == "logo":
                        # Handle the image data here
                        image_data = base64.b64decode(value)
                        timestamp = int(time.time())
                        unique_filename = f"logo_{timestamp}.png"
                        
                        owner.logo.save(unique_filename, ContentFile(image_data))
        
                    else:
                        setattr(owner, field, value)
                    
                owner.save()
                owner_data = {
                    "id": str(owner.id),
                    "business_name": owner.business_name,
                    "business_type": owner.business_type,
                    "first_name": owner.first_name,
                    "last_name": owner.last_name,
                    "email": owner.email,
                    "contact_no": owner.contact_no,
                    "address": owner.address,
                    "logo": owner.logo.url if owner.logo else None,
                    "tuition_tagline": owner.tuition_tagline if owner.tuition_tagline else None,
                    "status": owner.status,
                    "created_at": owner.created_at,
                    "city": {
                        "city_id": str(owner.city.id),
                        "city_name": owner.city.name,
                        "state_id": str(owner.city.state_id),
                        "state_name": owner.city.state.name,
                    },
                }

                response_data = {
                    "result": True,
                    "data": owner_data,
                    "message": "Profile updated successfully"
                }
            
                return response_data
            else:
                response_data = {
                        "result": False,
                        "message": "updation not apply"
                    }
                return JsonResponse(response_data, status=400)      
        else:
            response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
            return JsonResponse(response_data, status=400)
        

    except BusinessOwners.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Business owner not found"
        }
        return JsonResponse(response_data, status=400)
    except Cities.DoesNotExist:
        response_data = {
            "result": False,
            "message": "City not found"
        }
        return JsonResponse(response_data, status=400)
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=500)
    

#-----------------------------------------------------------------------------------------------------------#
#---------------------------------------------------NEWS----------------------------------------------------#
#-----------------------------------------------------------------------------------------------------------#

    
def add_news(data, user):
    try:
        owner = BusinessOwners.objects.get(id=user.id)

        # Check if both text and image are provided
        if data.text and data.image:
            response_data = {
                "result": False,
                "message": "You can only provide either text or an image, not both"
            }
            return JsonResponse(response_data, status=400)
        
        if data.batch and data.standard:
            response_data = {
                "result": False,
                "message": "You can only specify either 'batch' or 'standard', not both."
            }
            return JsonResponse(response_data, status=400)

            # Check if neither standard nor batch is provided
        
        # Check if neither text nor image is provided
        if not data.text and not data.image:
            response_data = {
                "result": False,
                "message": "You must provide either text or an image"
            }
            return JsonResponse(response_data, status=400)

        business_news = BusinessNewses(business_owner=user)

        if data.text:
            business_news.news = data.text
        if data.image:
            image_path = data.image
            with open(image_path, 'rb') as image_file:
                binary_data = image_file.read()
                logo_base64 = base64.b64encode(binary_data).decode('utf-8')
            image_name = os.path.basename(image_path)
            business_news.image = image_name
            business_news.image.save(image_name, ContentFile(base64.b64decode(logo_base64)))

        if data.standard:
            try:
                academic_standard = AcademicStandards.objects.get(id=data.standard)
                business_news.standard = academic_standard
            except AcademicStandards.DoesNotExist:
                response_data = {
                    "result": False,
                    "message": "The specified academic standard does not exist"
                }
                return JsonResponse(response_data, status=400)

        if data.batch:
            try:
                competitive_batch = CompetitiveBatches.objects.get(id=data.batch)
                business_news.batch = competitive_batch
            except CompetitiveBatches.DoesNotExist:
                response_data = {
                    "result": False,
                    "message": "The specified batch does not exist"
                }
                return JsonResponse(response_data, status=400)

        business_news.save()

        news_data = {
            "id": str(business_news.id),
            "image": business_news.image.url if business_news.image else None,
            "text": business_news.news if business_news.news else None,
            "status": business_news.status,
            "standard": business_news.standard.id if business_news.standard else None,
            "batch": business_news.batch.id if business_news.batch else None
        }
        response_data = {
            "result": True,
            "data": news_data,
            "message": "News added successfully"
        }
        return JsonResponse(response_data)

    except Exception as e:
        response_data = {
            "result": False,
            "message": "Something went wrong"
        }
        return JsonResponse(response_data, status=500)
    

def get_news_list(request):
    try:
        newses = BusinessNewses.objects.filter(business_owner=request.user, status="active").order_by('-created_at')
        newses_list = [
            {
                "id": str(news.id),
                "image": news.image.url if news.image else None,
                "text": news.news if news.news else None,
                "status":news.status,
                "standard":str(news.standard.id) if news.standard else None,
                "batch":str(news.batch.id) if news.batch else None,
            }
            for news in newses
        ]
        paginated_news, items_per_page = paginate_data(request, newses_list)

        return {
            "result": True,
            "data": list(paginated_news),
            "message": "Data retrieved successfully",   
            "pagination": {
                "page": paginated_news.number,
                "total_docs": paginated_news.paginator.count,
                "total_pages": paginated_news.paginator.num_pages,
                "per_page": items_per_page,
            },
        }


    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=500)

    
def get_news(news_id, user):
    try:
        business_news = BusinessNewses.objects.get(id=news_id)

        news_data = {
                "id": str(business_news.id),
                "image": business_news.image.url if business_news.image else None,
                "text": business_news.news if business_news.news else None,
                "standard":str(business_news.standard.id) if business_news.standard else None,
                "batch":str(business_news.batch.id) if business_news.batch else None,
                "status":business_news.status
            }
        response_data = {
            "result":True,
            "data": news_data,
            "message": "news retrieved successfully "
        }
        return response_data
    
    except BusinessNewses.DoesNotExist:
        response_data = {
            "result": False,
            "message": "News not found"
        }
        return JsonResponse(response_data, status=400)
    
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=500)


def update_news(news_id, data):
    try:
        business_news = BusinessNewses.objects.get(id=news_id)
        
        if business_news:
            update_data = {field: value for field, value in data.dict().items() if value is not None}
            if update_data:
                if 'text' in update_data and 'image' in update_data:
                    response_data = {
                        "result": False,
                        "message": "You can only provide either text or an image, not both"
                    }
                    return JsonResponse(response_data, status=400)
                
                if 'batch' in update_data and 'standard' in update_data:
                    response_data = {
                        "result": False,
                        "message": "You can only specify either 'batch' or 'standard', not both."
                    }
                    return JsonResponse(response_data, status=400)

                image_path = update_data.get('image')
                for field, value in update_data.items():
                    if field == "text":
                        business_news.image = None
                        business_news.news = value
                    elif field == "standard":
                        try:
                            academic_standard = AcademicStandards.objects.get(id=value)
                            business_news.standard = academic_standard
                        except AcademicStandards.DoesNotExist:
                            response_data = {
                                "result": False,
                                "message": "The specified academic standard does not exist"
                            }
                            return JsonResponse(response_data, status=400)
                        setattr(business_news, field, academic_standard)  # Assign the whole academic_standard instance
                    elif field == "batch":
                        try:
                            competitive_batch = CompetitiveBatches.objects.get(id=value)
                            business_news.batch = competitive_batch
                        except CompetitiveBatches.DoesNotExist:
                            response_data = {
                                "result": False,
                                "message": "The specified batch does not exist"
                            }
                            return JsonResponse(response_data, status=400)
                        setattr(business_news, field, competitive_batch)  # Assign the whole competitive_batch instance
                    else:
                        setattr(business_news, field, value)
                
                if image_path:
                    business_news.news = None
                    with open(image_path, 'rb') as image_file:
                        binary_data = image_file.read()
                        logo_base64 = base64.b64encode(binary_data).decode('utf-8')
                    image_name = os.path.basename(image_path)
                    business_news.image.save(image_name, ContentFile(base64.b64decode(logo_base64)))

                business_news.save()
                
                # Fetch the updated academic_standard and competitive_batch instances
                updated_academic_standard = business_news.standard
                updated_competitive_batch = business_news.batch
                
                news_data = {
                    "id": str(business_news.id),
                    "image": business_news.image.url if business_news.image else None,
                    "text": business_news.news if business_news.news else None,
                    "status": business_news.status,
                    "standard": {
                        "id": updated_academic_standard.id if updated_academic_standard else None,
                        "name": updated_academic_standard.standard if updated_academic_standard else None,
                        # Include other relevant fields here
                    },
                    "batch": {
                        "id": updated_competitive_batch.id if updated_competitive_batch else None,
                        "name": updated_competitive_batch.name if updated_competitive_batch else None,
                        # Include other relevant fields here
                    }
                }
                response_data = {
                    "result": True,
                    "data": news_data,
                    "message": "News updated successfully"
                }
                return JsonResponse(response_data, status=200)

    except BusinessNewses.DoesNotExist:
        response_data = {
            "result": False,
            "message": "News not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
            "result": False,
            "message": "Something went wrong"
        }
        return JsonResponse(response_data, status=500)




def delete_news(news_id):
    try:
        business_news = BusinessNewses.objects.get(id=news_id)
        business_news.delete()

        response_data = {
            "result":True,
            "message":"Data Deleted Successfully"}
        return response_data
    
    except BusinessNewses.DoesNotExist:
        response_data = {
                    "result": False,
                    "message": "Nwes not found"
                }
        return JsonResponse(response_data, status=400)
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)  



#-----------------------------------------------------------------------------------------------------------#
#-------------------------------------------COMPETITIVE BATCH-----------------------------------------------#
#-----------------------------------------------------------------------------------------------------------#


def add_batch(data, user):
    try:
        existing_batch = CompetitiveBatches.objects.filter(business_owner=user, batch_name=data.batch_name).first()
        if existing_batch:
            response_data = {
                        "result": False,
                        "message": "A batch with the same name already exists for this owner."
                    }
            return JsonResponse(response_data, status=400)
        
        batch = CompetitiveBatches(batch_name=data.batch_name, business_owner=user)
        batch.save()
        business_owner = BusinessOwners.objects.get(id=batch.business_owner_id)
        batch_data = {
            "id": str(batch.id),
            "batch_name": batch.batch_name,
            "business_owner_id": str(batch.business_owner_id),
            "business_owner_name": business_owner.business_name,
            "status": batch.status,
            "created_at": batch.created_at.strftime('%Y-%m-%dT%H:%M:%S.%fZ'),
            "updated_at": batch.updated_at.strftime('%Y-%m-%dT%H:%M:%S.%fZ'),
            
        }
        response_data = {
            "result": True,
            "data": batch_data,
            "message": "Batch created successfully",
        }
        return response_data
    except Exception as e:
        response_data = {
                "result": False,
                "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)


def get_batchlist(request, query):
    try:

        batches = CompetitiveBatches.objects.filter(business_owner=request.user).order_by('-created_at')
        chapters = CompetitiveChapters.objects.filter(subject_name__business_owner=request.user)

        if query.chapter_id:
            chapters = chapters.filter(id=query.chapter_id)
            if chapters.exists():
                chapter = chapters.first()
                batches = chapter.batches.all().order_by('-created_at')
        if query.status:
            batches = batches.filter(status=query.status)
        if query.search:
            search_terms = query.search.strip().split()  
            search_query = Q()  
            for term in search_terms:
                search_query |= Q(batch_name__icontains=term) | Q(status__icontains=term) 

            batches = batches.filter(search_query)
        business_owner = BusinessOwners.objects.get(id=request.user.id)
    
        batches_list = [
            {
                "id": str(batch.id),
                "batch_name": batch.batch_name,
                "business_owner_id": str(batch.business_owner_id),
                "business_owner_name": business_owner.business_name,
                "status": batch.status,
                "created_at": batch.created_at,
                "updated_at": batch.updated_at,
            }
            for batch in batches
        ]
        paginated_batches, items_per_page = paginate_data(request, batches_list)

        return {
            "result": True,
            "data": list(paginated_batches),
            "message": "Data retrieved successfully",   
            "pagination": {
                "page": paginated_batches.number,
                "total_docs": paginated_batches.paginator.count,
                "total_pages": paginated_batches.paginator.num_pages,
                "per_page": items_per_page,
            },
        }

    
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)


def get_batch(batch_id, user):
    try:
        batch = CompetitiveBatches.objects.get(id=batch_id)
        business_owner = BusinessOwners.objects.get(id=user.id)
        batch_data = {
                "id": str(batch.id),
                "batch_name": batch.batch_name,
                "business_owner_id": str(batch.business_owner_id),
                "business_owner_name": business_owner.business_name,
                "status": batch.status,
                "created_at": batch.created_at,
                "updated_at": batch.updated_at,
            }
        response_data = {
            "result": True,
            "data": batch_data,
            "message": "Competitive batch retrieved successfully"
        }

        return response_data
    
    except CompetitiveBatches.DoesNotExist:
        response_data = {
                    "result": False,
                    "message": "Batch not found"
                }
        return JsonResponse(response_data, status=400)
    
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)


def update_batch(batch_id, data, user):
    try:
        batch = CompetitiveBatches.objects.get(id=batch_id)
        if batch:
            existing_batch = CompetitiveBatches.objects.filter(business_owner=user, batch_name=data.batch_name).first()
            if existing_batch:
                response_data = {
                            "result": False,
                            "message": "A batch with the same name already exists for this owner."
                        }
                return JsonResponse(response_data, status=400)
            update_data = {field: value for field, value in data.dict().items() if value is not None}
            if update_data:
                for field, value in update_data.items():
                    setattr(batch, field, value)
                batch.save()
                business_owner = BusinessOwners.objects.get(id=batch.business_owner_id)
                response_data = {
                    "result": True,
                    "data": {
                        "id": str(batch.id),
                        "batch_name": batch.batch_name,
                        "business_owner_id": str(batch.business_owner_id),
                        "business_owner_name": business_owner.business_name,
                        "status": batch.status,
                        "created_at": batch.created_at.strftime('%Y-%m-%dT%H:%M:%S.%fZ'),
                        "updated_at": batch.updated_at.strftime('%Y-%m-%dT%H:%M:%S.%fZ'),
                    },
                    "message": "Competitive batch updated successfully",
                }
                return response_data
            else:
                response_data = {
                    "result": False,
                    "message": "No fields to update"
                }
                return JsonResponse(response_data, status=400)
                
        else:
            response_data = {
                    "result": False,
                    "message": "Batch not found"
                }
            return JsonResponse(response_data, status=400)
           
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)


def delete_batch(batch_id):
    try:
        batch = CompetitiveBatches.objects.get(id=batch_id)
        batch.delete()

        response_data = {
            "result":True,
            "message":"Data Deleted Successfully"}
        return response_data
    
    except CompetitiveBatches.DoesNotExist:
        response_data = {
                    "result": False,
                    "message": "Btach not found"
                }
        return JsonResponse(response_data, status=400)
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)  


#-----------------------------------------------------------------------------------------------------------#
#------------------------------------------COMPETITIVE SUBJECT----------------------------------------------#
#-----------------------------------------------------------------------------------------------------------#


def add_comp_subect(data, user):
    try:
        existing_subject = CompetitiveSubjects.objects.filter(business_owner=user, subject_name=data.subject_name).first()
        if existing_subject:
            response_data = {
                        "result": False,
                        "message": "A subject with the same name already exists for this owner."
                    }
            return JsonResponse(response_data, status=400)
        
        subject = CompetitiveSubjects.objects.create(subject_name=data.subject_name, business_owner=user)
        business_owner = BusinessOwners.objects.get(id=subject.business_owner_id)
        subject_data = {
            "id": str(subject.id),
            "subject_name": subject.subject_name,
            "business_owner_id": str(subject.business_owner_id),
            "business_owner_name": business_owner.business_name,
            "status": subject.status,
            "created_at": subject.created_at.strftime('%Y-%m-%dT%H:%M:%S.%fZ'),
            "updated_at": subject.updated_at.strftime('%Y-%m-%dT%H:%M:%S.%fZ'),
        }
        response_data = {
            "result": True,
            "data": subject_data,
            "message": "Subject created successfully",
        }
        return response_data

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)
    

def get_comp_subjectlist(request, query):
    try:
        subjects = CompetitiveSubjects.objects.filter(business_owner=request.user).order_by('-created_at')
        if query.status:
            subjects = subjects.filter(status=query.status)
        if query.search:
            search_terms = query.search.strip().split()  
            search_query = Q()  
            for term in search_terms:
                search_query |= Q(subject_name__icontains=term)  | Q(status__icontains=term)

            subjects = subjects.filter(search_query)
        business_owner = BusinessOwners.objects.get(id=request.user.id)
        subject_list = [
            {
                "id": str(subject.id),
                "subject_name": subject.subject_name,
                "business_owner_id": str(subject.business_owner_id),
                "business_owner_name": business_owner.business_name,
                "status": subject.status,
                "created_at": subject.created_at,
                "updated_at": subject.updated_at,
            }
            for subject in subjects
        ]
        paginated_comp_subjects, items_per_page = paginate_data(request, subject_list)

        return {
            "result": True,
            "data": list(paginated_comp_subjects),
            "message": "Data retrieved successfully",   
            "pagination": {
                "page": paginated_comp_subjects.number,
                "total_docs": paginated_comp_subjects.paginator.count,
                "total_pages": paginated_comp_subjects.paginator.num_pages,
                "per_page": items_per_page,
            },
        }

    
    except CompetitiveSubjects.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Subject not found"
        }
        return JsonResponse(response_data, status=400)
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)


def get_comp_subject(subject_id, user):
    try:
        subject = CompetitiveSubjects.objects.get(id=subject_id)
        business_owner = BusinessOwners.objects.get(id=user.id)
        subject_data = {
                "id": str(subject.id),
                "subject_name": subject.subject_name,
                "business_owner_id": str(subject.business_owner_id),
                "business_owner_name": business_owner.business_name,
                "status": subject.status,
                "created_at": subject.created_at,
                "updated_at": subject.updated_at,
            }
        response_data = {
            "result": True,
            "data": subject_data,
            "message": "Competitive subject retrieved successfully"
        }

        return response_data
    
    except CompetitiveSubjects.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Subject not found"
        }
        return JsonResponse(response_data, status=400)
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)

def update_comp_subject(subject_id, data, user):
    try:
        subject = CompetitiveSubjects.objects.get(id=subject_id)
        if subject:
            existing_subject = CompetitiveSubjects.objects.filter(business_owner=user, subject_name=data.subject_name).first()
            if existing_subject:
                response_data = {
                            "result": False,
                            "message": "A subject with the same name already exists for this owner."
                        }
                return JsonResponse(response_data, status=400)
            update_data = {field: value for field, value in data.dict().items() if value is not None}
            if update_data:
                for field, value in update_data.items():
                    setattr(subject, field, value)
                subject.save()
                business_owner = BusinessOwners.objects.get(id=subject.business_owner_id)
                response_data = {
                    "result": True,
                    "data": {
                        "id": str(subject.id),
                        "subject_name": subject.subject_name,
                        "business_owner_id": str(subject.business_owner_id),
                        "business_owner_name": business_owner.business_name,
                        "status": subject.status,
                        "created_at": subject.created_at.strftime('%Y-%m-%dT%H:%M:%S.%fZ'),
                        "updated_at": subject.updated_at.strftime('%Y-%m-%dT%H:%M:%S.%fZ'),
                    },
                    "message": "Subject updated successfully",
                }
                return response_data
            else:
                response_data = {
                    "result": False,
                    "message": "No fields to update"
                }
                return JsonResponse(response_data, status=400)
        else:
            response_data = {
                    "result": False,
                    "message": "Subject not found"
                }
            return JsonResponse(response_data, status=400)
        
    except CompetitiveSubjects.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Subject not found"
        }
        return JsonResponse(response_data, status=400)
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)


def delete_comp_subject(subject_id):
    try:
        subject = CompetitiveSubjects.objects.get(id=subject_id)
        subject.delete()

        response_data = {
            "result":True,
             "message":"Data Deleted Successfully"}
        return response_data

    except CompetitiveSubjects.DoesNotExist:
        response_data = {
                    "result": False,
                    "message": "Subject not found"
                }
        return JsonResponse(response_data, status=400)
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)
    

#-----------------------------------------------------------------------------------------------------------#
#------------------------------------------COMPETITIVE CHAPTER----------------------------------------------#
#-----------------------------------------------------------------------------------------------------------#


def add_comp_chapter(data, user):
    try:
        existing_chapter = CompetitiveChapters.objects.filter(subject_name__business_owner=user, chapter_name=data.chapter_name).first()
        if existing_chapter:
            response_data = {
                        "result": False,
                        "message": "A chapter with the same name already exists for this owner."
                    }
            return JsonResponse(response_data, status=400)

        subject_instance = CompetitiveSubjects.objects.get(id=data.subject_id)
        batches_instances = CompetitiveBatches.objects.filter(id__in=data.batches)
        batches = list(batches_instances)
        
        # Create a new chapter instance
        chapter = CompetitiveChapters(
            subject_name=subject_instance,
            chapter_name=data.chapter_name,
        )

        for batch in batches:
            chapter.batches.add(batch)
        # chapter_instance.batches = list(batches_instances)
        chapter.save()
        subject_id = CompetitiveSubjects.objects.get(id=chapter.subject_name_id)
        batch_info = [{"id": str(batch.id), "batch_name": batch.batch_name} for batch in chapter.batches.all()]
        chapter_data = {
            "id": str(chapter.id),
            "chapter_name": chapter.chapter_name,
            "subject_id": str(subject_id.id),
            "subject_name": subject_id.subject_name,
            "batches":batch_info,
            "status": chapter.status,
            "created_at": chapter.created_at,
            "updated_at": chapter.updated_at,
        }
        response_data = {
            "result": True,
            "data": chapter_data,
            "message": "Chapter created successfully",
        }
        return response_data
    
    except CompetitiveSubjects.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Subject not exist",
        }
        return JsonResponse(response_data, status=200)
    except Exception as e:
        response_data = {
            "result": False,
            "message": "Something went wrong",
        }
        return JsonResponse(response_data, status=200)


def get_comp_chapterlist(request, query):
    try:
        chapters = CompetitiveChapters.objects.filter(subject_name__business_owner=request.user).order_by('-created_at')

        if query.status:
            chapters = chapters.filter(status=query.status)
        if query.subject_id:
            chapters = chapters.filter(subject_name=query.subject_id)
        if query.batch_id:
            batch_id = query.batch_id
            chapters = [chapter for chapter in chapters if str(batch_id) in [str(batch.id) for batch in chapter.batches.all()]]
        if query.search:
            search_terms = query.search.strip().split()  
            search_query = Q()  
            for term in search_terms:
                search_query |= Q(chapter_name__icontains=term) | Q(subject_name__subject_name__icontains=term) | Q(batches__batch_name__icontains=term) | Q(status__icontains=term)

            chapters = chapters.filter(search_query)
        

        chapters_list = []
        if query.subject_ids:
            subject_ids_str = query.subject_ids.strip()
            subject_ids = subject_ids_str.split(",")
            
            for subject_id in subject_ids:
                chapters_for_subject = chapters.filter(subject_name__id=subject_id)
                subject_info = {
                    "subject_id": subject_id,
                    "subject_name": chapters_for_subject.first().subject_name.subject_name,
                    "chapters": [
                        {
                            "id": str(chapter.id),
                            "chapter_name": chapter.chapter_name,
                        }
                        for chapter in chapters_for_subject
                    ]
                }
                chapters_list.append(subject_info)
                paginated_comp_chapters, items_per_page = paginate_data(request, chapters_list)

            return {
                "result": True,
                "data": list(paginated_comp_chapters),
                "message": "Data retrieved successfully",   
                "pagination": {
                    "page": paginated_comp_chapters.number,
                    "total_docs": paginated_comp_chapters.paginator.count,
                    "total_pages": paginated_comp_chapters.paginator.num_pages,
                    "per_page": items_per_page,
                },
            }

        else:
            chapters = CompetitiveChapters.objects.all().order_by('-created_at')
            for chapter in chapters:
                subject_id = CompetitiveSubjects.objects.get(id=chapter.subject_name_id)
                batch_info = [{"id": str(batch.id), "batch_name": batch.batch_name} for batch in chapter.batches.all()]
                chapter_data = {
                    "id": str(chapter.id),
                    "chapter_name": chapter.chapter_name,
                    "subject_id": str(subject_id.id),
                    "subject_name": subject_id.subject_name,
                    "batches":batch_info,
                    "status": chapter.status,
                    "created_at": chapter.created_at,
                    "updated_at": chapter.updated_at,
                }
                
                chapters_list.append(chapter_data)
            paginated_comp_chapters, items_per_page = paginate_data(request, chapters_list)

            return {
                "result": True,
                "data": list(paginated_comp_chapters),
                "message": "Data retrieved successfully",   
                "pagination": {
                    "page": paginated_comp_chapters.number,
                    "total_docs": paginated_comp_chapters.paginator.count,
                    "total_pages": paginated_comp_chapters.paginator.num_pages,
                    "per_page": items_per_page,
                },
            }
    
    except CompetitiveChapters.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Chapter not found"
        }
        return JsonResponse(response_data, status=400)
    except CompetitiveSubjects.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Subject not found"
        }
        return JsonResponse(response_data, status=400)
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)

def get_comp_chapter(user, chapter_id):
    try:
        chapter = CompetitiveChapters.objects.get(id=chapter_id)
        subject_id = CompetitiveSubjects.objects.get(id=chapter.subject_name_id)
        batch_info = [{"id": str(batch.id), "batch_name": batch.batch_name} for batch in chapter.batches.all()]
        chapter_data = {
            "id": str(chapter.id),
            "chapter_name": chapter.chapter_name,
            "subject_id": str(subject_id.id),
            "subject_name": subject_id.subject_name,
            "batches":batch_info,
            "status": chapter.status,
            "created_at": chapter.created_at,
            "updated_at": chapter.updated_at,
        }
        response_data = {
                "result": True,
                "data": chapter_data,
                "message": "Competitive chapter retrieved successfully"
            }

        return response_data
    
    except CompetitiveChapters.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Chapter not found"
        }
        return JsonResponse(response_data, status=400)
    except CompetitiveSubjects.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Subject not found"
        }
        return JsonResponse(response_data, status=400)
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)


def update_comp_chapter(chapter_id, data, user):
    try:
        chapter = CompetitiveChapters.objects.get(id=chapter_id)
        if chapter:
            existing_chapter = CompetitiveChapters.objects.filter(subject_name__business_owner=user, chapter_name=data.chapter_name).first()
            if existing_chapter:
                response_data = {
                            "result": False,
                            "message": "A chapter with the same name already exists for this owner."
                        }
                return JsonResponse(response_data, status=400)
            update_data = {field: value for field, value in data.dict().items() if value is not None}
            if update_data:
                for field, value in update_data.items():

                    if field == "subject_id":  
                        subject = CompetitiveSubjects.objects.get(id=value)
                        chapter.subject_name = subject

                    elif field == "batches":  
                        new_batches_instances = CompetitiveBatches.objects.filter(id__in=value)
                        new_batches = list(new_batches_instances)
                        
                        # Get the existing batch IDs
                        existing_batch_ids = chapter.batches.all().values_list("id", flat=True)
                        
                        # Add new batches
                        for new_batch in new_batches:
                            if new_batch.id not in existing_batch_ids:
                                chapter.batches.add(new_batch)
                        
                        # Remove batches not in the updated list
                        for existing_batch_id in existing_batch_ids:
                            if existing_batch_id not in value:
                                batch_to_remove = CompetitiveBatches.objects.get(id=existing_batch_id)
                                chapter.batches.remove(batch_to_remove)

                    else:
                        setattr(chapter, field, value)

                chapter.save()

                subject_id = CompetitiveSubjects.objects.get(id=chapter.subject_name_id)
                batch_info = [{"id": str(batch.id), "batch_name": batch.batch_name} for batch in chapter.batches.all()]
                chapter_data = {
                    "id": str(chapter.id),
                    "chapter_name": chapter.chapter_name,
                    "subject_id": str(subject_id.id),
                    "subject_name": subject_id.subject_name,
                    "batches":batch_info,
                    "status": chapter.status,
                    "created_at": chapter.created_at,
                    "updated_at": chapter.updated_at,
                }
                response_data = {
                        "result": True,
                        "data": chapter_data,
                        "message": "Competitive chapter updated successfully"
                    }
                return response_data
    except CompetitiveChapters.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Chapter not found"
        }
        return JsonResponse(response_data, status=400)
    except CompetitiveSubjects.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Subject not found"
        }
        return JsonResponse(response_data, status=400)
    except CompetitiveBatches.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Batch not found"
        }
        return JsonResponse(response_data, status=400)
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400) 


def delete_comp_chapter(chapter_id):
    try:
        chapter = CompetitiveChapters.objects.get(id=chapter_id)
        chapter.delete()

        response_data = {"result":True,
                         "message":"Data Deleted Successfully"}
        return response_data
    
    except CompetitiveChapters.DoesNotExist:
        response_data = {
                    "result": False,
                    "message": "Chapter not found"
                }
        return JsonResponse(response_data, status=400)
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)


#-----------------------------------------------------------------------------------------------------------#
#-----------------------------------------COMPETITIVE QUESTIONS---------------------------------------------#
#-----------------------------------------------------------------------------------------------------------#


def add_comp_question(user,data):
    try:
        options_data = data.options  

        options_instance = Options.objects.create(
            option1=options_data.option1,
            option2=options_data.option2,
            option3=options_data.option3,
            option4=options_data.option4
        )

        competitive_chapter = CompetitiveChapters.objects.get(id= data.chapter_id)
        if data.question_image:
            # Handle the image data here
            image_data = base64.b64decode(data.question_image)
            timestamp = int(time.time())
            unique_filename = f"question_{timestamp}.png"
            competitive_question_image = ContentFile(image_data, name=unique_filename)
        else:
            competitive_question_image = None
        question = CompetitiveQuestions.objects.create(
            competitve_chapter=competitive_chapter,
            question=data.question,
            competitive_question_image=competitive_question_image,
            options=options_instance,  
            answer=data.answer,
            question_category=data.question_category,
            marks=data.marks,
            time_duration=data.time,
            business_owner=user
        )
        batch_info = [{"id": str(batch.id), "batch_name": batch.batch_name} for batch in competitive_chapter.batches.all()]
        question_data = {
            "id": str(question.id),
            "question": question.question,
            "question_image":question.competitive_question_image.url if question.competitive_question_image else None,
            "answer": question.answer,
            "options": options_data,  
            "chapter_id": str(question.competitve_chapter.id),
            "chapter_name": question.competitve_chapter.chapter_name,
            "subject_id": str(question.competitve_chapter.subject_name.id),
            "subject_name": question.competitve_chapter.subject_name.subject_name,
            "batches": batch_info,
            "question_category": question.question_category,
            "marks": str(question.marks),
            "time": str(question.time_duration),
            "status": question.status,
            "created_at":question.created_at,
            "updated_at":question.updated_at
        }

        response_data = {
            "result": True,
            "data": question_data,
            "message":"Question added successfully"
        }
        return response_data

    except Options.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Options not found"
        }
        return JsonResponse(response_data, status=400)
    
    except CompetitiveChapters.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Chapter not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)
 

def get_comp_questionlist(request, query):
    try:
        base_query = CompetitiveQuestions.objects.filter(business_owner=request.user).order_by('-created_at')
        search_query = Q()

        if query:
            if query.search:
                search_query |= (
                    Q(question__icontains=query.search)
                    | Q(answer__icontains=query.search)
                    | Q(competitve_chapter__chapter_name__icontains=query.search)
                    | Q(competitve_chapter__subject_name__subject_name__icontains=query.search)
                    | Q(competitve_chapter__batches__batch_name__icontains=query.search)
                    | Q(question_category__icontains=query.search)
                    | Q(marks__icontains=query.search)
                    | Q(time_duration__icontains=query.search)
                    | Q(status__icontains=query.search)
                )
            elif query.status:
                search_query &= Q(status=query.status)

            elif query.batch_id and query.chapter_id:
                search_query &= Q(competitve_chapter=query.chapter_id)

            elif query.subject_id and query.chapter_id:
                search_query &= Q(competitve_chapter=query.chapter_id)

            elif query.chapter_id:
                search_query &= Q(competitve_chapter=query.chapter_id)

            elif query.subject_id:
                search_query &= Q(competitve_chapter__subject_name=query.subject_id)

            elif query.question_category:
                search_query &= Q(question_category=query.question_category)

        questions = base_query.filter(search_query)
        batch_name_to_id = {}
        for question_data in questions.values('id', 'competitve_chapter__batches__batch_name'):
            batch_name = question_data['competitve_chapter__batches__batch_name']
            if batch_name not in batch_name_to_id:
                try:
                    batch = CompetitiveBatches.objects.get(batch_name=batch_name)
                    batch_name_to_id[batch_name] = str(batch.id)
                except CompetitiveBatches.DoesNotExist:
                    batch_name_to_id[batch_name] = None

        question_list = questions.values(
            'id', 'question', 'competitive_question_image', 'answer', 'marks', 'time_duration', 'status', 'created_at', 'updated_at', 'options_id', 'question_category',
            chapter_id=F('competitve_chapter__id'),
            chapter_name=F('competitve_chapter__chapter_name'),
            subject_id=F('competitve_chapter__subject_name__id'),
            subject_name=F('competitve_chapter__subject_name__subject_name'),
            batch_name=F('competitve_chapter__batches__batch_name')
        )

        options_data = Options.objects.filter(id__in=question_list.values_list('options_id', flat=True))
        options_dict = {option.id: option for option in options_data}

        # Add options to the question_list
        for question_data in question_list:
            option_id = question_data['options_id']
            options_data = options_dict.get(option_id)
            if options_data:
                question_data['options'] = {
                    'option1': options_data.option1,
                    'option2': options_data.option2,
                    'option3': options_data.option3,
                    'option4': options_data.option4,
                }
            else:
                question_data['options'] = {}

            batch_name = question_data['batch_name']
            question_data['batch_id'] = batch_name_to_id.get(batch_name)


        paginated_comp, items_per_page = paginate_data(request, question_list)

        return {
                "result": True,
                "data": list(paginated_comp),
                "message": "Data retrieved successfully",   
                "pagination": {
                    "page": paginated_comp.number,
                    "total_docs": paginated_comp.paginator.count,
                    "total_pages": paginated_comp.paginator.num_pages,
                    "per_page": items_per_page,
                },
            }

    except CompetitiveChapters.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Chapter not found"
        }
        return JsonResponse(response_data, status=400)

    except CompetitiveQuestions.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Question not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": str(e)
                }
        return JsonResponse(response_data, status=400)


def get_comp_question(user, question_id):
    try:
        question = CompetitiveQuestions.objects.get(id=question_id)
        try:
            options_data = Options.objects.get(id=question.options_id)
        except Options.DoesNotExist:
            options_data = None 
        
        options_dict = {
            "option1": options_data.option1 if options_data else None,
            "option2": options_data.option2 if options_data else None,
            "option3": options_data.option3 if options_data else None,
            "option4": options_data.option4 if options_data else None,
        }
        batch_info = [{"id": str(batch.id), "batch_name": batch.batch_name} for batch in question.competitve_chapter.batches.all()]
        question_data = {
                "id": str(question.id),
                "question": question.question,
                "question_image":question.competitive_question_image.url if question.competitive_question_image else None,
                "answer": question.answer,
                "options": options_dict, 
                "chapter_id": str(question.competitve_chapter.id),
                "chapter_name": question.competitve_chapter.chapter_name,
                "subject_id": str(question.competitve_chapter.subject_name.id),
                "subject_name": question.competitve_chapter.subject_name.subject_name, 
                "batches": batch_info,
                "question_category": question.question_category,
                "marks": str(question.marks),
                "time": str(question.time_duration),
                "status": question.status,
                "created_at":question.created_at,
                "updated_at":question.updated_at
            }
        
        response_data = {
            "result": True,
            "data": question_data,
            "message":"Question retrieved successfully"
        }
        return response_data
    except CompetitiveQuestions.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Question not found"
        }
        return JsonResponse(response_data, status=400)
    except Options.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Options not found"
        }
        return JsonResponse(response_data, status=400)
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)
    

def update_comp_question(question_id, data):
    try:
        question = CompetitiveQuestions.objects.get(id=question_id)
        try:
            options_data = Options.objects.get(id=question.options_id)
        except Options.DoesNotExist:
            options_data = None  
    
        update_data = {field: value for field, value in data.dict().items() if value is not None}
        
        if update_data:
            if "options" in update_data and options_data:
                new_options = update_data.pop("options")
                for field, value in new_options.items():
                    if hasattr(options_data, field) and value is not None:
                        setattr(options_data, field, value)
                options_data.save()

            if "question_image" in update_data and update_data["question_image"]:
                # Handle the image data here
                image_data = base64.b64decode(update_data["question_image"])
                timestamp = int(time.time())
                unique_filename = f"question_{timestamp}.png"
                competitive_question_image = ContentFile(image_data, name=unique_filename)
                question.competitive_question_image = competitive_question_image


            for field, value in update_data.items():
                if field == "chapter_id":
                    chapter = CompetitiveChapters.objects.get(id=value)
                    question.competitve_chapter = chapter
                if field == "time":
                    question.time_duration = value

                else:
                    setattr(question, field, value)
            question.save()
            batch_info = [{"id": str(batch.id), "batch_name": batch.batch_name} for batch in question.competitve_chapter.batches.all()]
            question_data = {
                "id": str(question.id),
                "question": question.question,
                "question_image":question.competitive_question_image.url if question.competitive_question_image else None,
                "answer": question.answer,
                "chapter_id": str(question.competitve_chapter.id),
                "chapter_name": question.competitve_chapter.chapter_name,
                "subject_id": str(question.competitve_chapter.subject_name.id),
                "subject_name": question.competitve_chapter.subject_name.subject_name, 
                "batches": batch_info,
                "question_category": question.question_category,
                "marks": str(question.marks),
                "time": str(question.time_duration),
                "status": question.status,
                "created_at":question.created_at,
                "updated_at":question.updated_at
            }
            if options_data:
                options_dict = {
                    "option1": options_data.option1,
                    "option2": options_data.option2,
                    "option3": options_data.option3,
                    "option4": options_data.option4,
                }
                question_data["options"] = options_dict
            else:
                options_dict = {
                    "option1": question_data.options_data.option1,
                    "option2": question_data.options_data.option2,
                    "option3": question_data.options_data.option3,
                    "option4": question_data.options_data.option4,
                }
                question_data["options"] = options_dict
        
            response_data = {
                "result": True,
                "data": question_data,
                "message":"Question retrieved successfully"
            }
            return response_data

    except CompetitiveQuestions.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Question not found"
        }
        return JsonResponse(response_data, status=400)
    except Options.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Options not found"
        }
        return JsonResponse(response_data, status=400)
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)
    

def delete_comp_question(question_id):
    try:
        question = CompetitiveQuestions.objects.get(id=question_id)
        options = Options.objects.get(id=question.options_id)
        options.delete()
        question.delete()
        

        response_data = {"result":True,
                         "message":"Data Deleted Successfully"}
        return response_data
    
    except CompetitiveQuestions.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Question not found"
        }
        return JsonResponse(response_data, status=400)
    except Options.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Options not found"
        }
        return JsonResponse(response_data, status=400)
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)


#-----------------------------------------------------------------------------------------------------------#
#--------------------------------------------COMPETITIVE EXAM-----------------------------------------------#
#-----------------------------------------------------------------------------------------------------------#


def create_comp_exam(user, data):
    try:
     
        batch_instance = CompetitiveBatches.objects.get(id=data.batch_id)
        total_weightage = data.total_questions
        selected_comp_questions_set1 = []
        selected_comp_questions_set2 = []        
        selected_comp_questions_set3 = [] 

        exam_data_calculated = []
        start_time = time.time()           
        def backtrack(selected_questions, remaining_time, remaining_marks, remaining_easy_questions, remaining_medium_questions, remaining_hard_questions, question_data_list):
            taken_time = time.time() - start_time
            if taken_time > 120:  # minutes in seconds
                raise TimeoutError("Questions not found.")
            if remaining_time < 0 and remaining_marks < 0:
                return False
            if remaining_easy_questions == 0 and remaining_medium_questions == 0 and remaining_hard_questions == 0:
                if remaining_time == 0 and remaining_marks == 0:
                    return True

            for _ in range(len(question_data_list)):
                selected_question = random.choice(question_data_list)
                if selected_question in question_data_list:
                    question_data_list.remove(selected_question)
                    
                if selected_question.question_category == "easy" and remaining_easy_questions > 0 and selected_question.time_duration <= remaining_time and selected_question.marks <= remaining_marks:
                    selected_questions.append(selected_question)
                    updated_remaining_time = remaining_time - selected_question.time_duration
                    updated_remaining_marks = remaining_marks - selected_question.marks
                    updated_remaining_easy_questions = remaining_easy_questions - 1
                    # print("new", updated_remaining_easy_questions)      
                    # print("ADSA", selected_questions)
                    # print("timeee", updated_remaining_time)
                    # print("markssss", updated_remaining_marks)
                    
                    if backtrack(selected_questions, updated_remaining_time, updated_remaining_marks, updated_remaining_easy_questions, remaining_medium_questions, remaining_hard_questions, question_data_list):
                        return True
                    else:
                        selected_questions.pop()
                        # print("SSSS")
                        return False
                    
                elif selected_question.question_category == "medium" and remaining_medium_questions > 0 and selected_question.time_duration <= remaining_time and selected_question.marks <= remaining_marks:
                    selected_questions.append(selected_question)
                    updated_remaining_time = remaining_time - selected_question.time_duration
                    updated_remaining_marks = remaining_marks - selected_question.marks
                    updated_remaining_medium_questions = remaining_medium_questions - 1
                    # print("new", updated_remaining_medium_questions)      
                    # print("ADSA", selected_questions)
                    # print("timeee", updated_remaining_time)
                    # print("markssss", updated_remaining_marks)
                    if backtrack(selected_questions, updated_remaining_time, updated_remaining_marks, remaining_easy_questions, updated_remaining_medium_questions, remaining_hard_questions, question_data_list):
                        return True
                    else:
                        selected_questions.pop()
                        # print("SSSS")
                        return False
                    
                elif selected_question.question_category == "hard" and remaining_hard_questions > 0 and selected_question.time_duration <= remaining_time and selected_question.marks <= remaining_marks:
                    selected_questions.append(selected_question)
                    updated_remaining_time = remaining_time - selected_question.time_duration
                    updated_remaining_marks = remaining_marks - selected_question.marks
                    updated_remaining_hard_questions = remaining_hard_questions - 1
                    # print("new", updated_remaining_hard_questions)      
                    # print("ADSA", selected_questions)
                    # print("timeee", updated_remaining_time)
                    # print("markssss", updated_remaining_marks)
                    if backtrack(selected_questions, updated_remaining_time, updated_remaining_marks, remaining_easy_questions, remaining_medium_questions, updated_remaining_hard_questions, question_data_list):
                        return True
                    else:
                        selected_questions.pop()
                        # print("SSSS")
                        return False

            return False

        subject_data_list = []
        
        for subject_data in data.exam_data:
            subject_weightage = sum(
                qtype for qtype in [subject_data.easy_question, subject_data.medium_question, subject_data.hard_question]
            )
            subject_percentage = round(subject_weightage / total_weightage, 2)

            subject_time = int(data.time_duration * subject_percentage + 0.5)
            subject_marks = round(float(data.total_marks * subject_percentage))
            print(subject_marks, "MMMMAAAARRRKKKKKK")
            print(subject_data.subject_id)
            print(subject_time)
            print(subject_marks)
            subject_instance = CompetitiveSubjects.objects.get(id=subject_data.subject_id)
            subject_data_list.append({
            "subject_id": subject_data.subject_id,
            "subject_time": subject_time,
            "subject_marks": subject_marks,
            # Include other relevant data if needed
        })
            # chapter_instance = CompetitiveChapters.objects.filter(id__in=subject_data.chapter)
            # chapters = list(chapter_instance)
            # chapter_ids = [f"{item.id}," for item in chapters]
            # chapters = " ".join(chapter_ids)
            # question_data = CompetitiveQuestions.objects.filter(competitve_chapter__subject_name=subject_instance)
            question_data = CompetitiveQuestions.objects.filter(competitve_chapter__id__in=subject_data.chapter)
            question_data_list = list(question_data)
            
            selected_questions_set1 = []
            backtrack_result_set1 = backtrack(selected_questions_set1, subject_time, subject_marks, subject_data.easy_question, subject_data.medium_question, subject_data.hard_question, question_data_list)
            
            selected_questions_set2 = []
            backtrack_result_set2 = backtrack(selected_questions_set2, subject_time, subject_marks, subject_data.easy_question, subject_data.medium_question, subject_data.hard_question, question_data_list)
            
            selected_questions_set3 = []
            backtrack_result_set3 = backtrack(selected_questions_set3, subject_time, subject_marks, subject_data.easy_question, subject_data.medium_question, subject_data.hard_question, question_data_list)
        

            while not backtrack_result_set1 and len(selected_questions_set1) != subject_weightage:
                question_data_list = list(question_data)  
                selected_questions_set1 = []  
                remaining_marks = subject_marks 
                
                remaining_time = subject_time 
                backtrack_result_set1 = backtrack(selected_questions_set1, remaining_time, remaining_marks, subject_data.easy_question, subject_data.medium_question, subject_data.hard_question, question_data_list)

            while not backtrack_result_set2 and len(selected_questions_set2) != subject_weightage:
                question_data_list = list(question_data)  
                selected_questions_set2 = []  
                remaining_marks = subject_marks 
                
                remaining_time = subject_time 
                backtrack_result_set2 = backtrack(selected_questions_set2, remaining_time, remaining_marks, subject_data.easy_question, subject_data.medium_question, subject_data.hard_question, question_data_list)

            while not backtrack_result_set3 and len(selected_questions_set3) != subject_weightage:
                question_data_list = list(question_data)  
                selected_questions_set3 = []  
                remaining_marks = subject_marks 
                
                remaining_time = subject_time
                backtrack_result_set3 = backtrack(selected_questions_set3, remaining_time, remaining_marks, subject_data.easy_question, subject_data.medium_question, subject_data.hard_question, question_data_list)

            if backtrack_result_set1:
                for question in selected_questions_set1:
                    options_data = Options.objects.get(id=question.options_id)
                    options_dict = {
                            "option1": options_data.option1 if options_data else None,
                            "option2": options_data.option2 if options_data else None,
                            "option3": options_data.option3 if options_data else None,
                            "option4": options_data.option4 if options_data else None,
                        }
                    comp_exam_instance = CompExam(
                        
                        id = str(question.id),
                        question=question.question,
                        time=float(question.time_duration),
                        mark=question.marks,
                        question_category=question.question_category,
                        subject = str(question.competitve_chapter),
                        options = options_dict,
                        answer = question.answer,
                        question_image = question.competitive_question_image.url if question.competitive_question_image else None,
                    )
                    selected_comp_questions_set1.append(comp_exam_instance)

            if backtrack_result_set2:
               
                for question in selected_questions_set2:
                    options_data = Options.objects.get(id=question.options_id)
                    options_dict = {
                            "option1": options_data.option1 if options_data else None,
                            "option2": options_data.option2 if options_data else None,
                            "option3": options_data.option3 if options_data else None,
                            "option4": options_data.option4 if options_data else None,
                        }
                    comp_exam_instance = CompExam(
                        id = str(question.id),
                        question=question.question,
                        time=float(question.time_duration),
                        mark=question.marks,
                        question_category=question.question_category,
                        subject = str(question.competitve_chapter),
                        options = options_dict,
                        answer = question.answer,
                        question_image = question.competitive_question_image.url if question.competitive_question_image else None
                    )
                    selected_comp_questions_set2.append(comp_exam_instance)

            if backtrack_result_set3:
                for question in selected_questions_set3:
                    options_data = Options.objects.get(id=question.options_id)
                    options_dict = {
                            "option1": options_data.option1 if options_data else None,
                            "option2": options_data.option2 if options_data else None,
                            "option3": options_data.option3 if options_data else None,
                            "option4": options_data.option4 if options_data else None,
                        }
                    comp_exam_instance = CompExam(
                        id = str(question.id),
                        question=question.question,
                        time=float(question.time_duration),
                        mark=question.marks,
                        question_category=question.question_category,
                        subject = str(question.competitve_chapter),
                        options = options_dict,
                        answer = question.answer,
                        question_image = question.competitive_question_image.url if question.competitive_question_image else None
                    )
                    selected_comp_questions_set3.append(comp_exam_instance)

            print("------------------------------------------------------")

        
        result = {
            # "exam_id": exam_instance.id,
            "subject_data": subject_data_list,
            "set1": selected_comp_questions_set1,
            "set2": selected_comp_questions_set2 if selected_comp_questions_set2 else None,
            "set3": selected_comp_questions_set3 if selected_comp_questions_set3 else None,
        }
      
        return result

    except TimeoutError as timeout_error:
        response_data = {
            "result": False,
            "message": "Backtracking took too long to complete"
        }
        return response_data
    
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)


def start_comp_CompExam(user, data):
    
    try:
        exam_data_calculated = []
        batch_id = data.batch_id  # Assuming 'standard_id' is a valid UUID4
        batch_instance = CompetitiveBatches.objects.get(id=batch_id)
        for subject_data in data.exam_data:
            easy=subject_data.easy_question
            medium=subject_data.medium_question
            hard=subject_data.hard_question
            subject_id = subject_data.subject_id
            subject_instance = CompetitiveSubjects.objects.get(id=subject_id)
            chapter_instance = CompetitiveChapters.objects.filter(id__in=subject_data.chapter)
            chapters = list(chapter_instance)
            chapter_ids = [f"{item.id}," for item in chapters]
            chapters = " ".join(chapter_ids)
        # Create a new exam object with the provided data
        exam = CompetitiveExams.objects.create(
            exam_title=data.exam_title,
            batch=batch_instance,
            total_questions=data.total_questions,
            time_duration=data.time_duration,
            passing_marks=data.passing_marks,
            total_marks=data.total_marks,
            negative_marks=data.negative_marks,
            option_e=data.option_e,
            business_owner=user,
  
        )

        # Get a list of selected question IDs from the provided data
        selected_question_ids = data.question

        # Filter and get the actual question instances
        selected_questions = CompetitiveQuestions.objects.filter(id__in=selected_question_ids)
        # print(selected_questions)
        # print(data.time)
        # print(data.mark)
        # Add the selected questions to the exam
        for question in selected_questions:
            exam.question_set.add(question)
        # Save the exam
    
        # print(f"Exam {exam.id} saved")

        subject_data = data.subject_data
        # print(subject_data)
        for subject_info in subject_data:
            subject_id = subject_info.subject_id
            subject_time = subject_info.subject_time
            subject_marks = subject_info.subject_marks
            subject_instance = CompetitiveSubjects.objects.get(id=subject_id)
            # ... (your existing code)

            exam_data_instance = CompetitiveExamData(
                subject=subject_instance,
                easy_question=easy,
                chapter=chapters,
                medium_question=medium,
                hard_question=hard,
                time_per_subject=subject_time,
                marks_per_subject=subject_marks,
            )
            exam_data_instance.save()
        exam_data_calculated.append(exam_data_instance)
        exam.save()
        for exam_data_instance in exam_data_calculated:
            exam.exam_data.add(exam_data_instance)  
  
        # for exam in exam_data_calculated:
        #     exam_instance.exam_data.add(exam) 
        result = {
            "result": True,
            "message": "Exam created and will start soon",
            "exam_id": exam.id  # Include the exam_id in the response
        }

        return JsonResponse(result) 
    
    except Exception as e:
        response_data = {
            "result": False,
            "message": "Something went wrong"
        }
        return JsonResponse(response_data, status=400)

def start_comp_exam(data):
    try:
        exam_id = data.exam_id  # Assuming 'exam_id' is in the payload
        print(exam_id)
        if not exam_id:
            response_data = {
                "result": False,
                "message": "exam_id is required in the payload"
            }
            return JsonResponse(response_data, status=400)

        exam = CompetitiveExams.objects.get(id=exam_id)
        print(exam,"thisgigfg")
        exam.start_date = datetime.now()
        exam.save()
        result = {
            "result": True,
            "message": "Exam started."
        }

        return result

    except AcademicExams.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Exam not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
            "result": False,
            "message": "Something went wrong"
        }
        return JsonResponse(response_data, status=400)


def get_comp_examlist(request, query):
    try: 
        exams = CompetitiveExams.objects.filter(business_owner=request.user, start_date__isnull=True).order_by('-created_at')
        exam_list = []
        
        if query.batch_id:
            exams = exams.filter(batch=query.batch_id)
        if query.subject_id:
            exams = exams.filter(exam_data__subject=query.subject_id)
        
        if query.chapter_id:
            exams = exams.filter(exam_data__chapter__contains=query.chapter_id)
        
        if query.search:
            search_terms = query.search.strip().split()
            search_query = Q()

            for term in search_terms:
                search_query |= (
                     Q(exam_title__icontains=term)
                    | Q(status__icontains=term)
                )
        for exam in exams:
            exam_data_list = []
            for exam_data in exam.exam_data.all():
                subject_name = exam_data.subject.subject_name
                chapter = exam_data.chapter
                # chapters = exam_data.chapter.split(",")

                # chapter_list = []
                # for chapter_id in chapters:
                #     if chapter_id:  
                #         print(chapter_id)
                #         chapter = CompetitiveChapters.objects.get(id=chapter_id)
                #         chapter_list.append({
                #             "chapter_name": chapter.chapter_name
                #         })
                exam_data_list.append({"subject": subject_name, "chapters": chapter})

            exam_detail = [
                {
                "id":str(exam.id),
                "exam_title": exam.exam_title,
                "batch": str(exam.batch.id),
                "batch_name": exam.batch.batch_name,
                "total_question":exam.total_questions,
                "time_duration":exam.time_duration,
                "negative_marks":exam.negative_marks,
                "total_marks":exam.total_marks,
                "start_date":exam.start_date,
                "exam_datas": exam_data_list
            }
            ]
            exam_list.append(exam_detail)
        paginated_comp_exam_data, items_per_page = paginate_data(request, exam_detail)

        return {
            "result": True,
            "data": list(paginated_comp_exam_data),
            "message": "Data retrieved successfully",   
            "pagination": {
                "page": paginated_comp_exam_data.number,
                "total_docs": paginated_comp_exam_data.paginator.count,
                "total_pages": paginated_comp_exam_data.paginator.num_pages,
                "per_page": items_per_page,
            },
        }
    
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)


#-----------------------------------------------------------------------------------------------------------#
#------------------------------------------------STUDENT----------------------------------------------------#
#-----------------------------------------------------------------------------------------------------------#


def create_student(data, user):
    try:
        existing_student = Students.objects.filter(business_owner=user, contact_no=data.contact_no).first()

        if existing_student:
            response_data = {
                "result": False,
                "message": "A student with this contact number already exists for this business owner."
            }
            return JsonResponse(response_data, status=400)

        existing_email = Students.objects.filter(business_owner=user, email=data.email).first()

        if existing_email:
            response_data = {
                "result": False,
                "message": "A student with this email already exists for this business owner."
            }
            return JsonResponse(response_data, status=400)
                
        if data.batch_id and data.standard_id:
            response_data = {
                "result": False,
                "message": "You can only specify either 'batch' or 'standard', not both."
            }
            return JsonResponse(response_data, status=400)

        student_data = {
            "business_owner": user,
            "first_name": data.first_name,
            "last_name": data.last_name,
            "email": data.email,
            "contact_no": data.contact_no,
            "parent_name": data.parent_name,
            "parent_contact_no": data.parent_contact_no,
            "profile_image": data.profile_image,
            "address": data.address,
        }
        
        if data.batch_id:
            batch_instance = CompetitiveBatches.objects.get(id=data.batch_id)
            student_data["batch"] = batch_instance
        
        if data.standard_id:
            standard_instance = AcademicStandards.objects.get(id=data.standard_id)
            student_data["standard"] = standard_instance

        student = Students.objects.create(**student_data)

        response_data = {
            "result": True,
            "data": {
                "id": str(student.id),
                "first_name": student.first_name,
                "last_name": student.last_name,
                "email": student.email,
                "contact_no": student.contact_no,
                "parent_name": student.parent_name,
                "parent_contact_no": student.parent_contact_no,
                "profile_image": student.profile_image.url if student.profile_image else None,
                "address": student.address,
                "competitive": {
                        "batch": str(student.batch.id) if student.batch else None,
                        "batch_name":student.batch.batch_name if student.batch else None
                },
                "academic":{
                        "board": str(student.standard.medium_name.board_name.id) if student.standard else None,
                        "board_name": student.standard.medium_name.board_name.board_name if student.standard else None,
                        "medium": str(student.standard.medium_name.id) if student.standard else None,
                        "medium_name": student.standard.medium_name.medium_name if student.standard else None,
                        "standard": str(student.standard.id) if student.standard else None,
                        "standard_name": student.standard.standard if student.standard else None  
                }
            },
            "message": "Student created successfully."
        }
        return response_data
    
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)
 

def upload_student(xl_file, user):
    try:
        xl_data = pd.read_excel(xl_file.file)
        created_students = []

        for _, row in xl_data.iterrows():
            student_info = {
                "first_name": row["first_name"],
                "last_name": row["last_name"],
                "email": row["email"],
                "contact_no": row["contact_no"],
                "parent_name": row["parent_name"],
                "parent_contact_no": row["parent_contact_no"],
                "address": row["address"],
            }

            standard_name = row.get("standard_name")  # Change to "standard_id"
            if standard_name:
                standard_instance = AcademicStandards.objects.get(standard=standard_name)
                student_info["standard"] = standard_instance

            batch_name = row.get("batch_name")  # Change to "batch_id"
            if batch_name:
                batch_instance = CompetitiveBatches.objects.get(batch_name=batch_name)  # Use batch_id
                student_info["batch"] = batch_instance
            
            existing_student = Students.objects.filter(business_owner=user, contact_no=student_info["contact_no"]).first()
            if existing_student:
                return JsonResponse({
                    "result": False,
                    "message": "Student with this contact number and email already exists."
                }, status=400)

            student_info["business_owner"] = user
            student = Students.objects.create(**student_info)
            created_students.append(student)
            
        response_data = {
            "result": True,
            "message": "Students created successfully."
        }
        return response_data

    except Exception as e:
        response_data = {
            "result": False,
            "message": "Something went wrong"
        }
        return JsonResponse(response_data, status=400)


def create_excel_with_column_names_student(file_path,flag,related_id, sheet_name="Sheet1"):

    try:
        timestamp = int(time.time())
        
        if flag == "standard":
            if related_id.standard_id:
                standard = AcademicStandards.objects.get(id=related_id.standard_id)
                standard_name = standard.standard
                print(standard_name)
            column_names = ["standard_name","first_name","last_name","email","contact_no","parent_name","parent_contact_no","address"] if related_id.standard_id else ["standard_name"]
    
        elif flag == "batch":
            if related_id.batch_id:
                batch = CompetitiveBatches.objects.get(id=related_id.batch_id)
                batch_name = batch.batch_name

            column_names = ["batch_name","first_name","last_name","email","contact_no","parent_name","parent_contact_no","address"] if related_id.batch_id else ["batch_name"]
    
        # Create a new workbook

       
        file_path = f"{timestamp}.xlsx"

        workbook = xlsxwriter.Workbook(file_path)
        worksheet = workbook.add_worksheet(sheet_name)

        for col_num, header in enumerate(column_names):
            worksheet.set_column(col_num, col_num, len(header) + 30)
        # Write column headers
        for col_num, header in enumerate(column_names):
            worksheet.write(0, col_num, header)
        
        if related_id.standard_id:
            for row_num in range(1, 30):  # Assuming 1000 rows for example
                worksheet.write(row_num, 0, standard_name)

        elif related_id.batch_id:
            for row_num in range(1, 30):  # Assuming 1000 rows for example
                worksheet.write(row_num, 0, batch_name)

        # Close the workbook
        workbook.close()

        with open(file_path, "rb") as file:
            file_content = file.read()
        
        response = HttpResponse(file_content, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=add_student_formate_{file_path}'
        return response
       
    except Exception as e:
        response_data = {
            "result": False,
            "message": "Something went wrong"
        }
        
        return response_data
# Example usage


def student_list(request, query):
    try:
        students = Students.objects.filter(business_owner=request.user).order_by('-created_at')
        if query.status:
            students = students.filter(status=query.status)
        elif query.board_id and query.medium_id and query.standard_id:
            students = students.filter(standard__id=query.standard_id)
        elif query.board_id and query.medium_id:
            students = students.filter(standard__medium_name__id=query.medium_id)
        elif query.batch_id:
            students = students.filter(batch__id=query.batch_id)
        elif query.board_id:
            students = students.filter(standard__medium_name__board_name__id=query.board_id)
        elif query.medium_id:
            students = students.filter(standard__medium_name__id=query.medium_id)
        elif query.standard_id:
            students = students.filter(standard__id=query.standard_id)
        elif query.search:
            search_terms = query.search.strip().split()
            search_query = Q()

            for term in search_terms:
                search_query |= (
                    Q(first_name__icontains=term)
                    | Q(last_name__icontains=term)
                    | Q(email__icontains=term)
                    | Q(parent_name__icontains=term)
                    | Q(address__icontains=term)
                    | Q(batch__batch_name__icontains=term)
                    | Q(standard__standard__icontains=term)
                    | Q(standard__medium_name__medium_name__icontains=term)
                    | Q(standard__medium_name__board_name__board_name__icontains=term)
                    | Q(status__icontains=term)
                )

            students = students.filter(search_query)
        else:
            students = Students.objects.filter(business_owner=request.user).order_by('-created_at')
        student_list = []
        for student in students:
            
            student_data = {
                "id": str(student.id),
                "first_name": student.first_name,
                "last_name": student.last_name,
                "email": student.email,
                "contact_no": student.contact_no,
                "parent_name": student.parent_name,
                "parent_contact_no": student.parent_contact_no,
                "profile_image": student.profile_image.url if student.profile_image else None,
                "address": student.address,
                "status":student.status,
                "competitive": {
                        "batch": str(student.batch.id) if student.batch else None,
                        "batch_name":student.batch.batch_name if student.batch else None
                },
                "academic":{
                        "board": str(student.standard.medium_name.board_name.id) if student.standard else None,
                        "board_name": student.standard.medium_name.board_name.board_name if student.standard else None,
                        "medium": str(student.standard.medium_name.id) if student.standard else None,
                        "medium_name": student.standard.medium_name.medium_name if student.standard else None,
                        "standard": str(student.standard.id) if student.standard else None,
                        "standard_name": student.standard.standard if student.standard else None  
                }
                
            }
            student_list.append(student_data)

        paginated_comp_exam_data, items_per_page = paginate_data(request, student_list)

        return {
            "result": True,
            "data": list(paginated_comp_exam_data),
            "message": "Data retrieved successfully",   
            "pagination": {
                "page": paginated_comp_exam_data.number,
                "total_docs": paginated_comp_exam_data.paginator.count,
                "total_pages": paginated_comp_exam_data.paginator.num_pages,
                "per_page": items_per_page,
            },
        }
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)


def student_detail(student_id):
    try:
        student = Students.objects.get(id=student_id)
        student_data = {
                "id": str(student.id),
                "first_name": student.first_name,
                "last_name": student.last_name,
                "email": student.email,
                "contact_no": student.contact_no,
                "parent_name": student.parent_name,
                "parent_contact_no": student.parent_contact_no,
                "profile_image": student.profile_image.url if student.profile_image else None,
                "address": student.address,
                "competitive": {
                        "batch": str(student.batch.id) if student.batch else None,
                        "batch_name":student.batch.batch_name if student.batch else None
                },
                "academic":{
                        "board": str(student.standard.medium_name.board_name.id) if student.standard else None,
                        "board_name": student.standard.medium_name.board_name.board_name if student.standard else None,
                        "medium": str(student.standard.medium_name.id) if student.standard else None,
                        "medium_name": student.standard.medium_name.medium_name if student.standard else None,
                        "standard": str(student.standard.id) if student.standard else None,
                        "standard_name": student.standard.standard if student.standard else None  
                }  
        }
        response_data = {
            "result": True,
            "data": student_data,
            "message": "Student retrieved successfully"
        }

        return response_data
    
    except Students.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Student not found"
        }
        return JsonResponse(response_data, status=400)
    
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": str(e)
                }
        return JsonResponse(response_data, status=400)


def student_updation(student_id, data):
    try:
        student = Students.objects.get(id=student_id)
        if student:
            existing_student = Students.objects.filter(business_owner=student.business_owner, contact_no=data.contact_no).first()

            if existing_student:
                response_data = {
                    "result": False,
                    "message": "A student with this contact number already exists for this business owner."
                }
                return JsonResponse(response_data, status=400)

            existing_email = Students.objects.filter(business_owner=student.business_owner, email=data.email).first()

            if existing_email:
                response_data = {
                    "result": False,
                    "message": "A student with this email already exists for this business owner."
                }
                return JsonResponse(response_data, status=400)
            
            if data.batch_id and student.standard_id:
                response_data = {
                    "result": False,
                    "message": "You cannot update 'batch' when 'standard' is already assigned."
                }
                return JsonResponse(response_data, status=400)

            if data.standard_id and student.batch_id:
                response_data = {
                    "result": False,
                    "message": "You cannot update 'standard' when 'batch' is already assigned."
                }
                return JsonResponse(response_data, status=400)

            update_data = {field: value for field, value in data.dict().items() if value is not None}
            if update_data:
                for field, value in update_data.items():
                    if field == "standard_id":
                        try:
                            standard_instance = AcademicStandards.objects.get(id=value)
                            student.standard = standard_instance
                        except AcademicStandards.DoesNotExist:
                            response_data = {
                                "result": False,
                                "message": "Invalid standard UUID provided."
                            }
                            return JsonResponse(response_data, status=400)

                    elif field == "batch_id":
                        try:
                            batch_instance = CompetitiveBatches.objects.get(id=value)
                            student.batch = batch_instance
                        except CompetitiveBatches.DoesNotExist:
                            response_data = {
                                "result": False,
                                "message": "Invalid batch UUID provided."
                            }
                            return JsonResponse(response_data, status=400)

                    else:
                        setattr(student, field, value)
                
                student.save()

                student_data = {
                    "id": str(student.id),
                    "first_name": student.first_name,
                    "last_name": student.last_name,
                    "email": student.email,
                    "contact_no": student.contact_no,
                    "parent_name": student.parent_name,
                    "parent_contact_no": student.parent_contact_no,
                    "profile_image": student.profile_image.url if student.profile_image else None,
                    "address": student.address,
                    "competitive": {
                        "batch": str(student.batch.id) if student.batch else None,
                        "batch_name":student.batch.batch_name if student.batch else None
                    },
                    "academic":{
                            "board": str(student.standard.medium_name.board_name.id) if student.standard else None,
                            "board_name": student.standard.medium_name.board_name.board_name if student.standard else None,
                            "medium": str(student.standard.medium_name.id) if student.standard else None,
                            "medium_name": student.standard.medium_name.medium_name if student.standard else None,
                            "standard": str(student.standard.id) if student.standard else None,
                            "standard_name": student.standard.standard if student.standard else None  
                    }  
                }
                response_data = {
                    "result": True,
                    "data": student_data,
                    "message": "Student updated successfully."
                }

                return response_data
    
    except Students.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Student not found."
        }
        return JsonResponse(response_data, status=404)
    except Exception as e:
        response_data = {
            "result": False,
            "message": "Something went wrong"
        }
        return JsonResponse(response_data, status=400)



def student_file_updation(xl_file, user):
    try:
        xl_data = pd.read_excel(xl_file.file)
        updated_students = []

        for _, row in xl_data.iterrows():
            student_info = {
                "contact_no": row["contact_no"]
            }
        
            student = Students.objects.filter(business_owner=user, contact_no=student_info["contact_no"]).first()
            batch_name = row.get("batch")
            if batch_name:
                batch_instance, _ = CompetitiveBatches.objects.get(batch_name=batch_name)
                student.batch = batch_instance
            
            board_name = row.get("board")
            if board_name:
                board_instance = AcademicBoards.objects.get(board_name=board_name)  

                medium_name = row.get("medium")
                if medium_name:
                    medium_instance = AcademicMediums.objects.get(board_name=board_instance, medium_name=medium_name)
                 
                    standard_id = row.get("standard")
                    if standard_id:
                        standard_instance = AcademicStandards.objects.get(medium_name=medium_instance, standard=standard_id)
                        student.standard = standard_instance


            student.save()
            updated_students.append(student)
        
        response_data = {
            "result": True,
            "message": "Students updated successfully."
        }
        return response_data
  
    except Exception as e:
        return JsonResponse({
            "result": False,
            "message": "Something went wrong"
        }, status=400)



def remove_student(student_id):
    try:
        student = Students.objects.get(id=student_id)
        student.delete()

        response_data = {"result":True,
                         "message":"Data Deleted Successfully"}
        return response_data
    
    except CompetitiveChapters.DoesNotExist:
        response_data = {
                    "result": False,
                    "message": "Chapter not found"
                }
        return JsonResponse(response_data, status=400)
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)


#-----------------------------------------------------------------------------------------------------------#
#-------------------------------------------------REPORT----------------------------------------------------#
#-----------------------------------------------------------------------------------------------------------#




def get_examreport(request, query):
    try:
        user = request.user
        if user.business_type == "competitive":
            exams = CompetitiveExams.objects.filter(business_owner=user.id, start_date__isnull=False)
            if query.batch_id:
                exams = exams.filter(batch=query.batch_id)
            if query.subject_id:
                exams = exams.filter(exam_data__subject=query.subject_id)
            if query.start_date and query.end_date:
                exams = exams.filter(start_date__gte=query.start_date, start_date__lte=query.end_date)
        elif user.business_type == "academic":
            exams = AcademicExams.objects.filter(business_owner=user.id, start_date__isnull=False)
            if query.board_id:
                exams = exams.filter(standard__medium_name__board_name=query.board_id)
            if query.medium_id:
                exams = exams.filter(standard__medium_name=query.medium_id)
            if query.standard_id:
                exams = exams.filter(standard=query.standard_id)
            if query.subject_id:
                exams = exams.filter(exam_data__subject=query.subject_id)
            if query.start_date and query.end_date:
                exams = exams.filter(start_date__gte=query.start_date, start_date__lte=query.end_date)
        else:
            return JsonResponse({
                "result": False,
                "message": "Invalid business type"
            }, status=400)

        exam_detail = []
        for exam in exams:
            results = Results.objects.filter(competitive_exam=exam.id) if user.business_type == "competitive" else Results.objects.filter(academic_exam=exam.id)
            passed_students = results.filter(result='pass').count()
            failed_students = results.filter(result='fail').count()
            total_students = results.count()

            exam_data = {
                "id": str(exam.id),
                "exam_title": exam.exam_title,
                "total_question": exam.total_questions,
                "time_duration": exam.time_duration,
                "negative_marks": exam.negative_marks,
                "total_marks": exam.total_marks,
                "start_date": exam.start_date,
                "total_students": total_students,
                "passed_students": passed_students,
                "failed_students": failed_students
            }

            if user.business_type == "academic":
                exam_data.update({
                    "board_id": str(exam.standard.medium_name.board_name.id),
                    "board_name": exam.standard.medium_name.board_name.board_name,
                    "medium_id": str(exam.standard.medium_name.id),
                    "medium_name": exam.standard.medium_name.medium_name,
                    "standard_id": str(exam.standard.id),
                    "standard_name": exam.standard.standard,
                })
            else:
                exam_data.update({
                    "batch": str(exam.batch.id),
                    "batch_name": exam.batch.batch_name,
                })

            exam_detail.append(exam_data)

        paginated_comp_exam_data, items_per_page = paginate_data(request, exam_detail)

        return {
            "result": True,
            "data": list(paginated_comp_exam_data) if paginated_comp_exam_data else [],
            "message": "Data retrieved successfully",   
            "pagination": {
                "page": paginated_comp_exam_data.number,
                "total_docs": paginated_comp_exam_data.paginator.count,
                "total_pages": paginated_comp_exam_data.paginator.num_pages,
                "per_page": items_per_page,
            },
        }
    except Exception as e:
        response_data = {
            "result": False,
            "message": "Something went wrong"
        }
        return JsonResponse(response_data, status=400)


def generate_pdf_report(exam_detail):
    buffer = BytesIO()
    timestamp = int(time.time())
    pdf = SimpleDocTemplate(buffer, pagesize=letter)
    content = []

    # Extract exam_info from exam_detail dictionary
    header_data = [
        [f"Exam Title: {exam_detail.get('exam_title', '')}",
        f"Total Questions: {exam_detail.get('total_question', '')}",
        f"Time Duration: {exam_detail.get('time_duration', '')}"],
        [''],
        [f"Negative Marks: {exam_detail.get('negative_marks', '')}",
        f"Total Marks: {exam_detail.get('total_marks', '')}",
        ],
        [''],
        [f"Passed_students: {exam_detail.get('passed_students','')}",
         f"failed_students: {exam_detail.get('failed_students','')}",
         f"total_students:{exam_detail.get('total_students','')}"],
        [''],
        [f"Start Date: {exam_detail.get('start_date', '')}"],
    ]
    if 'board_id' in exam_detail:
        header_data.append([''])
        header_data.append([f"Board Name: {exam_detail.get('board_name', '')}",
                            f"Medium Name: {exam_detail.get('medium_name', '')}",
                            f"Standard Name: {exam_detail.get('standard_name', '')}"
                            ])
        header_data.append([''])
    elif 'batch' in exam_detail:
        header_data.append([''])
        header_data.append([f"Batch Name: {exam_detail.get('batch_name', '')}"])
        header_data.append([''])
    header_table = Table(header_data, colWidths=[2.3*inch, 2.3*inch, 2.3*inch])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),  # Set header background color
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),  # Set header text color
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),  # Use a bold font
        ('FONTSIZE', (0, 0), (-1, -1), 12),  # Set font size
    ]))
    content.append(header_table)

    # Extract students_data from exam_detail dictionary
    students_data = exam_detail.get("students_data", [])
    student_info = [['First Name', 'Last Name', 'Right Answers', 'Wrong Answers', 'Mark', 'Time']]
    for student in students_data:
        student_info.append([
            student.get('student_first_name', ''),
            student.get('student_last_name', ''),
            str(student.get('total_right_answers', '')),
            str(student.get('total_wrong_answers', '')),
            str(student.get('mark', '')),
            str(student.get('time', ''))
        ])

    students_table = Table(student_info, colWidths=[1.5*inch, 1.5*inch, 1*inch, 1*inch, 1*inch, 1*inch])
    students_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2ecc71')),  # Set table header background color
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),  # Set table header text color
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Add grid lines
    ]))
    content.append(students_table)

    pdf.build(content)
    buffer.seek(0)
    filename = f"exam_report_{timestamp}.pdf"
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    buffer.close()
    return response



def exam_detail_report(request, exam_id, query):
    try:
        user_type = request.user.business_type
        if user_type not in ["competitive", "academic"]:
            raise ValueError("Invalid business type")
        
        exam_model = CompetitiveExams if user_type == "competitive" else AcademicExams
        result_model = Results.objects.filter(competitive_exam=exam_id) if user_type == "competitive" else Results.objects.filter(academic_exam=exam_id)
        passed_students = result_model.filter(result='pass').count()
        failed_students = result_model.filter(result='fail').count()
        total_students = result_model.count()
        exam = exam_model.objects.get(id=exam_id)
        
        exam_detail = {
                "exam_title": exam.exam_title,
                "total_question": exam.total_questions,
                "time_duration": exam.time_duration,
                "negative_marks": exam.negative_marks,
                "total_marks": exam.total_marks,
                "start_date": exam.start_date,
                "passed_students": passed_students,
                "failed_students": failed_students,
                "total_students": total_students
        }
        if request.user.business_type == "academic":
            exam_detail.update({
                "board_id": str(exam.standard.medium_name.board_name.id),
                "board_name": exam.standard.medium_name.board_name.board_name,
                "medium_id": str(exam.standard.medium_name.id),
                "medium_name": exam.standard.medium_name.medium_name,
                "standard_id": str(exam.standard.id),
                "standard_name": exam.standard.standard,
            })
        else:
            exam_detail.update({
                "batch": str(exam.batch.id),
                "batch_name": exam.batch.batch_name,
            })
        students_data = []
        for result in result_model:
            student = Students.objects.get(id=result.student.id)
            wrong_answers = StudentAnswers.objects.filter(
                student=student.id,
                competitive_exam=exam_id if user_type == "competitive" else None,
                academic_exam=exam_id if user_type == "academic" else None,
                is_correct__in=[False]
            ).count()
            right_answers = StudentAnswers.objects.filter(
                student=student.id,
                competitive_exam=exam_id if user_type == "competitive" else None,
                academic_exam=exam_id if user_type == "academic" else None,
                is_correct__in=[True]
            ).count()

            student_data = {
                    "student_first_name": student.first_name,
                    "student_last_name": student.last_name,
                    "total_right_answers": right_answers,
                    "total_wrong_answers": wrong_answers,
                    "mark": result.score,
                    "time": result.time_duration
            }
            students_data.append(student_data)

        exam_detail["students_data"] = students_data

        if query.generate_pdf:
            pdf_response = generate_pdf_report(exam_detail)
            return pdf_response

        response_data = {
            "result": True,
            "data": exam_detail,
            "message":"Report retrived successfully"
        }
        return response_data

    except Exception as e:
        response_data = {
            "result": False,
            "message": str(e)
        }
        return JsonResponse(response_data, status=400)































































































































































































































































































































































































































































































































####################################################################################
####--------------------------------ACADEMIC------------------------------------####
####################################################################################


def get_boards_list(request,filter_prompt):
    try:
        academic_boards = AcademicBoards.objects.filter(business_owner=request.user).order_by('-created_at')

        if filter_prompt.search:
            q_objects = Q(board_name__icontains=filter_prompt.search) | Q(business_owner__business_name__icontains=filter_prompt.search)
            academic_boards = academic_boards.filter(q_objects)
            
        elif filter_prompt.status:
            academic_boards = academic_boards.filter(status=filter_prompt.status)
        # else:
        #     academic_boards = AcademicBoards.objects.all()
        
        academic_list = [
            {
                "id": str(board.id),
                "board_name": board.board_name,
                "business_owner": board.business_owner.business_name,
                "status": board.status,
                "created_at": board.created_at,
                "updated_at": board.updated_at,
            }
            for board in academic_boards
        ]
        paginated_boards, items_per_page = paginate_data(request, academic_list)

        return {
            "result": True,
            "data": list(paginated_boards),
            "message": "Data retrieved successfully",   
            "pagination": {
                "page": paginated_boards.number,
                "total_docs": paginated_boards.paginator.count,
                "total_pages": paginated_boards.paginator.num_pages,
                "per_page": items_per_page,
            },
        }
    

    except AcademicBoards.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Board not found"
        }
        return JsonResponse(response_data, status=400)
    
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)
    
    
def get_academic_board_data(user,board_id):
    try:
        academic_boards = AcademicBoards.objects.get(id=board_id)

        academic_board = {
                "id": str(academic_boards.id),
                "board_name": academic_boards.board_name,
                "business_owner": academic_boards.business_owner.business_name,
                "status": academic_boards.status,
                "created_at": academic_boards.created_at,
                "updated_at": academic_boards.updated_at,
            }

        response_data = {
            "result": True,
            "data": academic_board,
            "message": "Academic board retrieved successfully."
        }

        return response_data
    
    except AcademicBoards.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Board not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)


def add_baord(user, data):
    try:
        # Create a new AcademicBoard instance
        board = AcademicBoards.objects.create(
            board_name=data.board_name,
            business_owner=user,
        )
        # Return the saved board data as a response
        saved_board = {
            "id": str(board.id),
            "board_name": board.board_name,
            "business_owner": user.business_name,  # Use the business owner's name from the token
            "status": board.status,
            "created_at": board.created_at,
            "updated_at": board.updated_at,
        }

        response_data = {
            "result": True,
            "data": saved_board,
            "message": "board added successfully."
        }

        return response_data

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)


def create_board(board_name, user):
    board_instance, created = AcademicBoards.objects.get_or_create(
        board_name=board_name,
        business_owner=user,
    )
    return board_instance, created


def create_medium(medium_name, board_id):
    board = AcademicBoards.objects.get(pk=board_id)
    medium_instance, created = AcademicMediums.objects.get_or_create(
        medium_name=medium_name,
        board_name=board,
    )
    return medium_instance, created

def create_standard(standard_name, medium_id):
    medium = AcademicMediums.objects.get(pk=medium_id)
    standard_instance, created = AcademicStandards.objects.get_or_create(
        standard=standard_name,
        medium_name=medium,
    )
    return standard_instance, created

def create_subject(subject_name, standard_id):
    standard = AcademicStandards.objects.get(pk=standard_id)
    subject_instance, created = AcademicSubjects.objects.get_or_create(
        subject_name=subject_name,
        standard=standard,
    )
    return subject_instance, created

def create_chapter(chapter_name, subject_id):
    print(f"Creating chapter: {chapter_name} for subject ID: {subject_id}")
    subject = AcademicSubjects.objects.get(pk=subject_id)
    chapter_instance, created = AcademicChapters.objects.get_or_create(
        chapter_name=chapter_name,
        subject_name=subject,
    )
    return chapter_instance, created

def create_question(academic_chapter, question, options, answer, question_category, marks, time_duration,img_data, business_owner):
    if img_data:
            # Handle the image data here
            image_data = base64.b64decode(img_data)
            timestamp = int(time.time())
            unique_filename = f"question_{timestamp}.png"
            academic_question_image = ContentFile(image_data, name=unique_filename)
    else:
        academic_question_image = None
    question_instance, created = AcademicQuestions.objects.get_or_create(
        academic_chapter=academic_chapter,
        question=question,
        options=options,
        answer=answer,
        question_category=question_category,
        marks=marks,
        time_duration=time_duration,
        business_owner=business_owner,
        academic_question_image=academic_question_image
    )
    return question_instance, created

def create_batch(batch_name, user):
    batch_instance, created = CompetitiveBatches.objects.get_or_create(
        batch_name=batch_name,
        business_owner=user,
    )
    return batch_instance, created

def create_competitive_subject(subject_name, user):
    subject_instance, created = CompetitiveSubjects.objects.get_or_create(
        subject_name=subject_name,
        business_owner=user
    )
    return subject_instance, created

def create_competitive_chapter(chapter_name, subject_instance, batch_ids):
    print(batch_ids)
    print(subject_instance)

    competitive_chapter_instance, created = CompetitiveChapters.objects.get_or_create(
        chapter_name=chapter_name,
        subject_name=subject_instance
    )
    
    for batch_id in batch_ids:
        batch_instance = CompetitiveBatches.objects.get(id=batch_id)
        competitive_chapter_instance.batches.add(batch_instance)

    return competitive_chapter_instance, created

def create_competitive_question(competitive_chapter, question, options, answer, question_category, marks, time_duration, img_data,business_owner):

    if img_data:
            # Handle the image data here
            image_data = base64.b64decode(img_data)
            timestamp = int(time.time())
            unique_filename = f"question_{timestamp}.png"
            competitive_question_image = ContentFile(image_data, name=unique_filename)
    else:
        competitive_question_image = None
    question_instance, created = CompetitiveQuestions.objects.get_or_create(
        competitve_chapter=competitive_chapter,
        question=question,
        options=options,
        answer=answer,
        question_category=question_category,
        marks=marks,
        time_duration=time_duration,
        business_owner=business_owner,
        competitive_question_image=competitive_question_image
    )
    return question_instance, created

def get_options_image_data(column_name, Option1_start_row, xl_file):
    print(Option1_start_row, "start row")
    bucket_name = os.getenv('AWS_STORAGE_BUCKET_NAME')
    # Load the OpenPyXL workbook
    pxl_doc = openpyxl.load_workbook(xl_file.file)

    # Get the sheet containing the image
    sheet = pxl_doc['Sheet1']

    # Create an image loader object
    image_loader = SheetImageLoader(sheet)

    row_number = Option1_start_row

    while True:
        try:
            cell_content = sheet[f'{column_name}{row_number}'].value
            print(cell_content, "at cell content")
            if isinstance(cell_content, str):
                # If the cell contains text, yield it
                yield cell_content
            else:
                image = image_loader.get(f'{column_name}{row_number}')
                print(image)

                if image:
                    # Save the image to BytesIO
                    image_stream = BytesIO()
                    image.save(image_stream, format=image.format)
                    image_stream.seek(0)
                    timestamp = datetime.now().strftime("%Y%m%d%H%M%S%f")
                    # Upload the image to S3
                    s3_key = f"images/{column_name}_{row_number}_{timestamp}.{image.format}"
                    s3_client.upload_fileobj(image_stream, bucket_name, s3_key)

                    # Yield the S3 URL
                    s3_url = f"https://{bucket_name}.s3.amazonaws.com/{s3_key}"
                    yield s3_url
                else:
                    return None
        except Exception as e:
            return None
        
def get_image_data(column_name, start_row,xl_file):
    print(start_row,"start row")
    # Load the OpenPyXL workbook
    pxl_doc = openpyxl.load_workbook(xl_file.file)

    # Get the sheet containing the image
    sheet = pxl_doc['Sheet1']

    # Create an image loader object
    image_loader = SheetImageLoader(sheet)

    row_number = start_row

    while True:
        try:
            # Get the image object
            image = image_loader.get(f'{column_name}{row_number}')
            
            if image:
                # Get the image format
                image_format = image.format

                # Save the image in the original format
                image.save('image.{}'.format(image_format), format=image_format)

                # Encode the binary data from the image as base64
                with open('image.{}'.format(image_format), 'rb') as image_file:
                    binary_data = image_file.read()
                    base64_encoded_data = base64.b64encode(binary_data).decode('utf-8')

                yield base64_encoded_data
                
            else:
                return None
            # row_number += 1
        except Exception as e:
            return None

start_row = 2
academic_start_row = 2
Option1_start_row = 2
Option2_start_row = 2
Option3_start_row = 2
Option4_start_row = 2
aca_Option1_start_row = 2
aca_Option2_start_row = 2
aca_Option3_start_row = 2
aca_Option4_start_row = 2
def upload_from_xl(xl_file, user, flag,param_prompt):
    try:
        xl_data = pd.read_excel(xl_file.file)
        created_boards = []
        existing_boards = []
        created_mediums = []
        existing_mediums = []
        created_standards = []
        existing_standards = []
        created_subjects = []
        existing_subjects = []
        created_chapters = []
        existing_chapters = []
        created_questions = []
        existing_questions = []
        created_batches = []  
        existing_batches = []
        created_competitive_subjects = []  
        existing_competitive_subjects = [] 
        created_competitive_chapters= []  
        existing_competitive_chapters = [] 
        created_competitive_questions = []
        existing_competitive_questions = []
        error_messages = []
        error_messages2 = []
        for _, row in xl_data.iterrows():
            if flag == "board":
                board_name = row.get("board_name")
                if board_name:
                    existing_board = AcademicBoards.objects.filter(board_name=board_name).first()
                    if existing_board:
                        existing_boards.append(board_name)
                    else:
                        board_instance, created = create_board(board_name, user)
                        if created:
                            created_boards.append(board_instance)

            

            elif flag == "medium":
                medium_name = row.get("medium_name")
                board_id = row.get("board_id")
                if medium_name:
                    existing_medium = AcademicMediums.objects.filter(medium_name=medium_name).first()
                    if existing_medium:
                        existing_mediums.append(medium_name)
                    else:
                        medium_instance, created = create_medium(medium_name, board_id)
                        if created:
                            created_mediums.append(medium_instance)

            elif flag == "standard":
                standard_name = row.get("standard_name")
                medium_id = row.get("medium_id")
                if standard_name :
                    existing_standard = AcademicStandards.objects.filter(standard=standard_name).first()
                    if existing_standard:
                        existing_standards.append(standard_name)
                    else:
                        standard_instance, created = create_standard(standard_name,medium_id)
                        if created:
                            created_standards.append(standard_instance)

            elif flag == "subject":
                subject_name = row.get("subject_name")
                standard_id = row.get ("standard_id")
                if subject_name:
                    existing_subject = AcademicSubjects.objects.filter(subject_name=subject_name).first()
                    if existing_subject:
                        existing_subjects.append(subject_name)
                    else:
                        subject_instance, created = create_subject(subject_name,standard_id)
                        if created:
                            created_subjects.append(subject_instance)
                        

            elif flag == "chapter":
                chapter_name = row.get("chapter_name")
                subject_id = row.get("subject_id")
                if chapter_name:
                    existing_chapter = AcademicChapters.objects.filter(chapter_name=chapter_name).first()
                    if existing_chapter:
                        existing_chapters.append(chapter_name)
                    else:
                        chapter_instance, created = create_chapter(chapter_name,subject_id)
                        if created:
                            created_chapters.append(chapter_instance)

            elif flag == "academic_question":
                aca_column_name = "G"
                aca_option1_column_name = 'H'
                aca_option2_column_name = 'I'
                aca_option3_column_name = 'J'
                aca_option4_column_name = 'K'
                board_name = row.get("board_name")
                medium_name = row.get("medium_name")
                standard_name = row.get("standard_name")
                subject_name = row.get("subject_name")
                chapter_name = row.get("chapter_name")
                question_text = row.get("question")
                option1 = row.get("option1")
                option2 = row.get("option2")
                option3 = row.get("option3")
                option4 = row.get("option4")
                answer = row.get("answer")
                question_category = row.get("question_category")
                marks = row.get("marks")
                time_duration = row.get("time_duration")
                global academic_start_row, aca_Option1_start_row,aca_Option2_start_row,aca_Option3_start_row,aca_Option4_start_row

                image_generator = get_image_data(aca_column_name, academic_start_row,xl_file)
                print(image_generator)
                if image_generator is not None:
                    img_data = next(image_generator, None)
                academic_start_row +=1
                
                if (
                    board_name and medium_name and standard_name and subject_name and
                    chapter_name and question_text and option1 and option2 and option3 and
                    option4 and answer and question_category and marks and time_duration
                ):
                    # Find all academic chapters that match the criteria
                    academic_chapters = AcademicChapters.objects.filter(
                        subject_name__subject_name=subject_name,
                        subject_name__standard__standard=standard_name,
                        subject_name__standard__medium_name__medium_name=medium_name,
                        subject_name__standard__medium_name__board_name__board_name=board_name,
                        chapter_name=chapter_name,
                        subject_name__standard__medium_name__board_name__business_owner=user
                    )

                    if not academic_chapters:
                        error_messages.append("No matching academic chapters found.")
                    else:
                        for academic_chapter in academic_chapters:
                            existing_question = AcademicQuestions.objects.filter(
                                question=question_text,
                                academic_chapter=academic_chapter
                            ).first()

                            if existing_question:
                                existing_questions.append(question_text)
                            else:
                                image_option_1 = get_options_image_data(aca_option1_column_name, aca_Option1_start_row,xl_file)
                                if image_option_1 is not None:
                                    option1_data = next(image_option_1, None)
                                    print(option1_data,"this is at option1 ")
                                aca_Option1_start_row += 1

                                image_option_2 = get_options_image_data(aca_option2_column_name, aca_Option2_start_row,xl_file)
                                if image_option_2 is not None:
                                    option2_data = next(image_option_2, None)
                                    print(option2_data,"this is at option2 ")
                                aca_Option2_start_row += 1

                                image_option_3 = get_options_image_data(aca_option3_column_name, aca_Option3_start_row,xl_file)
                                if image_option_3 is not None:
                                    option3_data = next(image_option_3, None)
                                    print(option3_data,"this is at option3 ")
                                aca_Option3_start_row += 1

                                image_option_4 = get_options_image_data(aca_option4_column_name, aca_Option4_start_row,xl_file)
                                if image_option_4 is not None:
                                    option4_data = next(image_option_4, None)
                                    print(option4_data,"this is at option4 ")
                                aca_Option4_start_row += 1
                                # Create options instance
                                options_instance = Options.objects.create(
                                    option1=option1_data,
                                    option2=option2_data,
                                    option3=option3_data,
                                    option4=option4_data
                                )

                                # Create question instance
                                question_instance, created = create_question(
                                    academic_chapter, question_text, options_instance, answer, question_category, marks, time_duration,img_data, user
                                )

                                if created:
                                    created_questions.append(question_instance)
                        
                else:
                    error_messages2.append("some fileds are missing.")
                


            elif flag == "batch":
                batch_name = row.get("batch_name")

                if batch_name:
                    existing_batch = CompetitiveBatches.objects.filter(batch_name=batch_name).first()
                    if existing_batch:
                        existing_batches.append(batch_name)
                    else:
                        batch_instance, created = create_batch(batch_name, user)
                        if created:
                            created_batches.append(batch_instance)


            elif flag == "competitive_subject":
                subject_name = row.get("subject_name")
        
                if subject_name:
                    existing_subject = CompetitiveSubjects.objects.filter(subject_name=subject_name).first()
                    if existing_subject:
                        existing_competitive_subjects.append(subject_name)
                    else:
                        subject_instance, created = create_competitive_subject(subject_name, user)
                        if created:
                            created_competitive_subjects.append(subject_instance)   

            elif flag == "competitive_chapter":
                chapter_name = row.get("chapter_name")
                subject_id = row.get("subject_ids")  # Assuming you have a field for subject_id in your Excel file
                batch_ids = row.get("batch_ids")    # Assuming you have a field for batch_ids in your Excel file

                if chapter_name and subject_id and batch_ids:
                    existing_chapter = CompetitiveChapters.objects.filter(
                        chapter_name=chapter_name,
                    ).first()

                    if existing_chapter:
                        existing_competitive_chapters.append(chapter_name)
                    else:
                        # Create a list to hold batch instances
                        batch_instances = []

                        for batch_id in batch_ids.split(','):
                            batch_instances.append(batch_id)

                        try:
                            subject_instance = CompetitiveSubjects.objects.get(id=subject_id)
                            print(subject_instance)
                            # print(f"subject_id type: {type(subject_id)}, subject_id value: {subject_id}")

                        except Exception as e:
                            # Handle other exceptions
                            print(f"Error: {e}")

                        # Create competitive chapter instance
                        competitive_chapter_instance, created = create_competitive_chapter(
                            chapter_name, subject_instance, batch_instances
                        )

                        if created:
                            created_competitive_chapters.append(competitive_chapter_instance)


            elif flag == "competitive_question":
                column_name = 'E'  # Example column name
                option1_column_name = 'F'
                option2_column_name = 'G'
                option3_column_name = 'H'
                option4_column_name = 'I'
                batch_name = row.get("batch_name")
                subject_name = row.get("subject_name")
                chapter_name = row.get("chapter_name")
                question_text = row.get("question")
                option1 = row.get("option1")
                option2 = row.get("option2")
                option3 = row.get("option3")
                option4 = row.get("option4")
                answer = row.get("answer")
                question_category = row.get("question_category")
                marks = row.get("marks")
                time_duration = row.get("time_duration")
                global start_row, Option1_start_row,Option2_start_row,Option3_start_row,Option4_start_row
         
                image_generator = get_image_data(column_name, start_row,xl_file)
            
                if image_generator is not None:
                    img_data = next(image_generator, None)
    
                start_row += 1             
                if (
                    batch_name and subject_name and
                    chapter_name and question_text and option1 and option2 and option3 and
                    option4 and answer and question_category and marks and time_duration
                ):

                    competitive_chapters = CompetitiveChapters.objects.filter(
                        chapter_name=chapter_name,
                        subject_name__business_owner = user
                    )
                    if not competitive_chapters:
                        error_messages.append("No matching competitive chapters found.")
                    else:
                        for competitive_chapter in competitive_chapters:
                            existing_question = CompetitiveQuestions.objects.filter(
                                question=question_text,
                                competitve_chapter=competitive_chapter
                            ).first()

                            if existing_question:
                                existing_competitive_questions.append(question_text)
                                # print(existing_competitive_questions)
                            else:
                                image_option_1 = get_options_image_data(option1_column_name, Option1_start_row,xl_file)
                                if image_option_1 is not None:
                                    option1_data = next(image_option_1, None)
                                    print(option1_data,"this is at option1 ")
                                Option1_start_row += 1

                                image_option_2 = get_options_image_data(option2_column_name, Option2_start_row,xl_file)
                                if image_option_2 is not None:
                                    option2_data = next(image_option_2, None)
                                    print(option2_data,"this is at option2 ")
                                Option2_start_row += 1

                                image_option_3 = get_options_image_data(option3_column_name, Option3_start_row,xl_file)
                                if image_option_3 is not None:
                                    option3_data = next(image_option_3, None)
                                    print(option3_data,"this is at option3 ")
                                Option3_start_row += 1

                                image_option_4 = get_options_image_data(option4_column_name, Option4_start_row,xl_file)
                                if image_option_4 is not None:
                                    option4_data = next(image_option_4, None)
                                    print(option4_data,"this is at option4 ")
                                Option4_start_row += 1

                                
                                options_instance = Options.objects.create(
                                    
                                    option1=option1_data,
                                    option2=option2_data,
                                    option3=option3_data,
                                    option4=option4_data
                                )
                                print(options_instance)
                                # Create question instance
                                question_instance, created = create_competitive_question(
                                    competitive_chapter, question_text, options_instance, answer, question_category, marks, time_duration,img_data,user
                                )

                                if created:
                                    created_competitive_questions.append(question_instance)
                                
                else:
                    error_messages2.append("some fileds are missing.")
        
        academic_start_row = 2
        start_row = 2 
        Option1_start_row = 2 
        Option2_start_row = 2
        Option3_start_row = 2
        Option4_start_row = 2  
        aca_Option1_start_row = 2
        aca_Option2_start_row = 2
        aca_Option3_start_row = 2
        aca_Option4_start_row = 2
                            
        response_data = {
            "result": True,
            "message": "Data uploaded successfully.",
     
        }

        if existing_boards:
            response_data["message"] = "Some boards already exist."

        if existing_mediums:
            response_data["message"] = "Some mediums already exist."

        if existing_standards:
            response_data["message"] = "Some standards already exist."
        
        if existing_subjects:
            response_data["message"] = "Some subjects already exist."
        
        if existing_chapters:
            response_data["message"] = "Some chapters already exist."
        
        if existing_questions:
            response_data["message"] = "Some questions already exist."

        if existing_batches:
            response_data["message"] = "Some batches already exist."
        
        if existing_competitive_subjects:
            response_data["message"] = "Some competitive subjects already exist."

        if existing_competitive_chapters:
            response_data["message"] = "Some competitive chapters already exist."

        if existing_competitive_questions:
            response_data["message"] = "Some questions already exist."

        if error_messages:
            response_data["message"] = "chapter not found for this user."
        
        if error_messages2:
            response_data["message"] = "some fileds are missing."
        return response_data

    except Exception as e:
        response_data = {
            "result": False,
            "message": "Something went wrong"
        }


        
import xlsxwriter
from fastapi.responses import JSONResponse
from django.http import HttpResponse
def create_excel_with_column_names(file_path,flag,related_id_name, sheet_name="Sheet1"):
    try:
        timestamp = int(time.time())
        if flag == "board":
            column_names = ["board_name"]
        elif flag == "medium":
            column_names = ["board_id", "medium_name"] if related_id_name.board_id else ["plase enter board_id in params"]
        elif flag == "standard":
            column_names = ["medium_id","standard_name"] if related_id_name.medium_id else ["plase enter medium_id in params"]
        elif flag == "subject":
            column_names = ["standard_id","subject_name"] if related_id_name.standard_id else ["plase enter standard_id in params"]
        elif flag == "chapter":
            column_names = ["subject_id","chapter_name"] if related_id_name.subject_id  else ["plase enter subject_id in params"]
        elif flag == "question":
            column_names = ["chapter_id","question","option1","option2","option3","option4","answer","question_category","marks","time_duration"] if related_id_name.chapter_id else ["plase enter chapter_id in params"]
        elif flag == "batch":
            column_names = ["batch_name"]
        elif flag == "competitive_subject":
            column_names = ["subject_name"]
        elif flag == "competitive_chapter":
            column_names = ["subject_id","batch_ids","chapter_name"] if related_id_name.subject_ids and related_id_name.batch_ids else ["plase enter subject_ids and batch_ids in params"]
        elif flag == "competitive_question":
            column_names = ["chapter_id","question","option1","option2","option3","option4","answer","question_category","marks","time_duration"] if related_id_name.chapter_id else ["plase enter chapter_id in params"]
        # Create a new workbook
        elif flag == "academic_subject":
            if related_id_name.subject_id:
                subject = AcademicSubjects.objects.get(id=related_id_name.subject_id)
                chapters_related_to_subject = AcademicChapters.objects.filter(subject_name=related_id_name.subject_id)
                subject_name = subject.subject_name
                standard_name = subject.standard.standard
                medium_name = subject.standard.medium_name.medium_name
                board_name = subject.standard.medium_name.board_name.board_name
            column_names = ["board_name","medium_name","standard_name","subject_name","chapter_name","question","option1","option2","option3","option4","answer","question_category","marks","time_duration"] if standard_name else ["plase enter standard_id in params"]
        
        elif flag == "competitve_subject":
            if related_id_name.competitive_subject_id:
                subject = CompetitiveSubjects.objects.get(id=related_id_name.competitive_subject_id)
                chapters_related_to_subject = CompetitiveChapters.objects.filter(subject_name=related_id_name.competitive_subject_id)
                if chapters_related_to_subject.exists():
                    # Get all batch names for the chapters_related_to_subject
                    batch_info = []
                    for chapter in chapters_related_to_subject:
                        batch_info.extend([{"batch_name": batch.batch_name} for batch in chapter.batches.all()])
                column_names = ["batch_name","subject_name","chapter_name","question","competitive_question_image","option1","option2","option3","option4","answer","question_category","marks","time_duration"]
        if flag == "competitve_subject":
            file_path = f"competitve_questions_format_{timestamp}.xlsx"
        else:
            file_path = f"academic_questions_format_{timestamp}.xlsx"

        workbook = xlsxwriter.Workbook(file_path)
        worksheet = workbook.add_worksheet(sheet_name)

        for col_num, header in enumerate(column_names):
            worksheet.set_column(col_num, col_num, len(header) + 30)

        # Write column headers
        for col_num, header in enumerate(column_names):
            worksheet.write(0, col_num, header)
        
        if related_id_name.board_id:
            for row_num in range(1, 30):  # Assuming 1000 rows for example
                worksheet.write(row_num, 0, related_id_name.board_id)

        elif related_id_name.medium_id:
            for row_num in range(1, 30):  # Assuming 1000 rows for example
                worksheet.write(row_num, 0, related_id_name.medium_id)
        
        elif related_id_name.standard_id:
            for row_num in range(1, 30):  # Assuming 1000 rows for example
                worksheet.write(row_num, 0, related_id_name.standard_id)

        
        # elif related_id_name.subject_id:
        #     for row_num in range(1, 30):  # Assuming 1000 rows for example
        #         worksheet.write(row_num, 0, related_id_name.subject_id)

        elif related_id_name.subject_id:
            for row_num, chapter in enumerate(chapters_related_to_subject, start=1):
                worksheet.write(row_num, 3, subject_name)
                worksheet.write(row_num, 2, standard_name)
                worksheet.write(row_num, 1, medium_name)
                worksheet.write(row_num, 0, board_name)
                worksheet.write(row_num, 4, chapter.chapter_name)

        elif related_id_name.competitive_subject_id:
            for row_num, chapter in enumerate(chapters_related_to_subject, start=1):
                for batch in batch_info:
                    worksheet.write(row_num, 0, batch["batch_name"])
                    worksheet.write(row_num, 1, subject.subject_name)
                    worksheet.write(row_num, 2, chapter.chapter_name)


        elif related_id_name.chapter_id:
            for row_num in range(1, 200):  # Assuming 1000 rows for example
                worksheet.write(row_num, 0, related_id_name.chapter_id)

        elif related_id_name.subject_ids and related_id_name.batch_ids:
            for row_num in range(1, 30):
                worksheet.write(row_num, 0, related_id_name.subject_ids)
                worksheet.write(row_num, 1, related_id_name.batch_ids)
                print(f"Row {row_num}: Subject ID = {related_id_name.subject_id}, Batch IDs = {related_id_name.batch_ids}")
        # Close the workbook
        workbook.close()

        with open(file_path, "rb") as file:
            file_content = file.read()

        # Return the Excel file as a response
        response = HttpResponse(file_content, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={file_path}'

        return response
       
    except Exception as e:
        response_data = {
            "result": False,
            "message": "Something went wrong"
        }
        
        return response_data
# Example usage


    
def update_board_data(user,data,board_id):
    try:
        # Check if the academic board exists
        academic_board = AcademicBoards.objects.get(id=board_id, business_owner=user)
        update_data = {field: value for field, value in data.dict().items() if value is not None}
        
        # Update the board_name and status
        for field, value in update_data.items():
            setattr(academic_board, field, value)
        
        academic_board.save()

        updated_board = {
            "id": str(academic_board.id),
            "board_name": academic_board.board_name,
            "business_owner": academic_board.business_owner.business_name,
            "status": academic_board.status,
            "created_at": academic_board.created_at,
            "updated_at": academic_board.updated_at,
        }

        response_data = {
            "result": True,
            "data": updated_board,
            "message": "Academic board updated successfully."
        }

        return response_data 
    
    except AcademicBoards.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Board not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)


def delete_board_data(user,board_id):
    try:
        # Check if the academic board exists
        academic_board = AcademicBoards.objects.get(id=board_id, business_owner=user)
        
        # Delete the academic board
        academic_board.delete()

        response_data = {
            "result": True,
            "message": "Academic board deleted successfully."
        }

        return response_data

    except AcademicBoards.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Board not found."
        }
        return JsonResponse(response_data, status=404)
    
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)
    

#-----------------------------------------------------------------------------------------------------------#
#-------------------------------------------------MEDIUM----------------------------------------------------#
#-----------------------------------------------------------------------------------------------------------#
   

def get_academic_mediums_list(request, filter_prompt):
    try:
        academic_mediums = AcademicMediums.objects.filter(board_name__business_owner=request.user).order_by('-created_at')

        if filter_prompt.search:
            q_objects = Q(medium_name__icontains=filter_prompt.search) | Q(board_name__board_name__icontains=filter_prompt.search)
            academic_mediums = academic_mediums.filter(q_objects)

        elif filter_prompt.status and filter_prompt.board_id and filter_prompt.medium_id:
            academic_mediums = academic_mediums.filter(
                status=filter_prompt.status,
                board_name__id=filter_prompt.board_id,
                id=filter_prompt.medium_id
            )
        elif filter_prompt.status and filter_prompt.board_id:
            academic_mediums = academic_mediums.filter(
                status=filter_prompt.status,
                board_name__id=filter_prompt.board_id
            )
        elif filter_prompt.status and filter_prompt.medium_id:
            academic_mediums = academic_mediums.filter(
                status=filter_prompt.status,
                id=filter_prompt.medium_id
            )
        elif filter_prompt.board_id and filter_prompt.medium_id:
            academic_mediums = academic_mediums.filter(
                board_name__id=filter_prompt.board_id,
                id=filter_prompt.medium_id
            )
        elif filter_prompt.status:
            academic_mediums = academic_mediums.filter(status=filter_prompt.status)
        elif filter_prompt.board_id:
            academic_mediums = academic_mediums.filter(board_name__id=filter_prompt.board_id)
        elif filter_prompt.medium_id:
            academic_mediums = academic_mediums.filter(id=filter_prompt.medium_id)
        academic_medium_list = [
            {
                "id": str(medium.id),
                "medium_name": medium.medium_name,
                "board_id": str(medium.board_name.id),
                "board_name": medium.board_name.board_name,
                "status": medium.status,
                "created_at": medium.created_at,
                "updated_at": medium.updated_at,
            }
            for medium in academic_mediums
        ]
        paginated_mediums, items_per_page = paginate_data(request, academic_medium_list)

        return {
            "result": True,
            "data": list(paginated_mediums),
            "message": "Data retrieved successfully",   
            "pagination": {
                "page": paginated_mediums.number,
                "total_docs": paginated_mediums.paginator.count,
                "total_pages": paginated_mediums.paginator.num_pages,
                "per_page": items_per_page,
            },
        }
    
    except AcademicMediums.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Medium not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)
    

def get_academic_medium_data(user,medium_id):
    try:
        academic_mediums = AcademicMediums.objects.get(id=medium_id)

        academic_medium_list = {
                "id": str(academic_mediums.id),
                "medium_name": academic_mediums.medium_name,
                "board_id": str(academic_mediums.board_name.id),
                "board_name": academic_mediums.board_name.board_name,
                "status": academic_mediums.status,
                "created_at": academic_mediums.created_at,
                "updated_at": academic_mediums.updated_at,
            }
    
        response_data = {
            "result": True,
            "data": academic_medium_list,
            "message": "Academic mediums retrieved successfully."
        }
        return response_data
    
    except AcademicMediums.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Medium not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)



def add_medium_data(user, data):
    try:
        board_id = data.get('board_id')  # Fetch board_id (UUID) from the request data
        board_instance = AcademicBoards.objects.get(id=board_id) 
        
        medium = AcademicMediums.objects.create(
            medium_name=data.get('medium_name'),
            board_name=board_instance,  # Use the fetched board instance
        )
    
        
        saved_medium = {
            "id": str(medium.id),
            "medium_name": medium.medium_name,
            "board_id": str(medium.board_name.id),
            "board_name": medium.board_name.board_name,
            "status": medium.status,
            "created_at": medium.created_at,
            "updated_at": medium.updated_at,
        }

        response_data = {
            "result": True,
            "data": saved_medium,
            "message": "Medium added successfully."
        }

        return response_data
    
    except AcademicBoards.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Board not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)



def update_medium_data(user, data, medium_id):
    try:
        # Check if the academic medium exists
        academic_medium = AcademicMediums.objects.get(id=medium_id)

        # Check if a new board_id is provided
        if data.board_id:
            try:
                academic_board = AcademicBoards.objects.get(id=data.board_id)
                academic_medium.board_name = academic_board
            except AcademicBoards.DoesNotExist:
                response_data = {
                    "result": False,
                    "message": "The specified academic board does not exist"
                }
                return JsonResponse(response_data, status=400)

        # Update the medium_name and status
        update_data = {field: value for field, value in data.dict().items() if value is not None}
        for field, value in update_data.items():
            setattr(academic_medium, field, value)

        academic_medium.save()

        updated_medium = {
            "id": str(academic_medium.id),
            "medium_name": academic_medium.medium_name,
            "board_id": str(academic_medium.board_name.id),
            "board_name": str(academic_medium.board_name.board_name),
            "status": academic_medium.status,
            "created_at": academic_medium.created_at,
            "updated_at": academic_medium.updated_at,
        }

        response_data = {
            "result": True,
            "data": updated_medium,
            "message": "Academic medium updated successfully."
        }

        return JsonResponse(response_data)

    except AcademicMediums.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Medium not found"
        }
        return JsonResponse(response_data, status=400)
    
    except AcademicBoards.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Board not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
            "result": False,
            "message": "Something went wrong"
        }
        return JsonResponse(response_data, status=500)


def delete_medium_data(user,medium_id):
    try:
        # Check if the academic board exists
        academic_medium = AcademicMediums.objects.get(id=medium_id)
        
        # Delete the academic board
        academic_medium.delete()

        response_data = {
            "result": True,
            "message": "Academic medium deleted successfully."
        }

        return response_data
    
    except AcademicMediums.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Medium not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)


#-----------------------------------------------------------------------------------------------------------#
#-------------------------------------------------STANDARD--------------------------------------------------#
#-----------------------------------------------------------------------------------------------------------#
    

def get_academic_standard_list(request, filter_prompt):
    try:
        academic_standards = AcademicStandards.objects.filter(medium_name__board_name__business_owner=request.user).order_by('-created_at')
        

        if filter_prompt.search:
            q_objects = (
                Q(medium_name__medium_name__icontains=filter_prompt.search) |
                Q(standard__icontains=filter_prompt.search)
            )
            academic_standards = academic_standards.filter(q_objects)

        elif filter_prompt.medium_id and filter_prompt.board_id:
            academic_standards = academic_standards.filter(medium_name__id=filter_prompt.medium_id)
        elif filter_prompt.status:
            academic_standards = academic_standards.filter(status=filter_prompt.status)
        elif filter_prompt.medium_id:
            academic_standards = academic_standards.filter(medium_name__id=filter_prompt.medium_id)
        elif filter_prompt.board_id:
            academic_standards = academic_standards.filter(medium_name__board_name__id=filter_prompt.board_id)
        elif filter_prompt.standard_id:
            academic_standards = academic_standards.filter(id=filter_prompt.standard_id)

        else:
            academic_standards = AcademicStandards.objects.filter(medium_name__board_name__business_owner=request.user)
        academic_standard_list = [
            {
                "id": str(standards.id),
                "standard": standards.standard,
                "medium_id": str(standards.medium_name.id),
                "medium_name": standards.medium_name.medium_name,
                "board_id": str(standards.medium_name.board_name.id),
                "board_name": str(standards.medium_name.board_name.board_name),
                "status": standards.status,
                "created_at": standards.created_at,
                "updated_at": standards.updated_at,
            }
            for standards in academic_standards
        ]
        
        paginated_standards, items_per_page = paginate_data(request, academic_standard_list)

        return {
            "result": True,
            "data": list(paginated_standards),
            "message": "Data retrieved successfully",   
            "pagination": {
                "page": paginated_standards.number,
                "total_docs": paginated_standards.paginator.count,
                "total_pages": paginated_standards.paginator.num_pages,
                "per_page": items_per_page,
            },
        }
    
    
    except AcademicStandards.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Standard not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)
    

def add_standard_data(user, data):

    try:

        medium_id = data.medium_id 
    
        medium_instance = AcademicMediums.objects.get(id=medium_id) 
     # Convert to UUID and fetch the corresponding board instance
        standards = AcademicStandards.objects.create(
            standard=data.standard,
            medium_name=medium_instance,  # Use the fetched board instance
        )
    
        
        saved_standard = {
            "id": str(standards.id),
            "standard":standards.standard,
            "medium_id":str(standards.medium_name.id),
            "medium_name": standards.medium_name.medium_name,
            "status": standards.status,
            "created_at": standards.created_at,
            "updated_at": standards.updated_at,
        }
     
        response_data = {
            "result": True,
            "data": saved_standard,
            "message": "standard added successfully."
        }

        return response_data 
    
    except AcademicMediums.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Medium not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400, safe=False)
    

def delete_standard_data(user,standard_id):
    try:
        # Check if the academic board exists
        academic_standard = AcademicStandards.objects.get(id=standard_id)
        
        # Delete the academic board
        academic_standard.delete()

        response_data = {
            "result": True,
            "message": "Academic standard deleted successfully."
        }

        return response_data
    
    except AcademicStandards.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Standard not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)
    

def get_academic_standard_data(standard_id):
    try:
        academic_standards = AcademicStandards.objects.get(id=standard_id)

        academic_standard_list = {
                "id": str(academic_standards.id),
                "standard":academic_standards.standard,
                "medium_id":str(academic_standards.medium_name.id),
                "medium_name": academic_standards.medium_name.medium_name,
                "board_id": str(academic_standards.medium_name.board_name.id),
                "board_name": str(academic_standards.medium_name.board_name.board_name),
                "status": academic_standards.status,
                "created_at": academic_standards.created_at,
                "updated_at": academic_standards.updated_at,
            }

        response_data = {
            "result": True,
            "data": academic_standard_list,
            "message": "Academic mediums retrieved successfully."
        }

        return response_data
    
    except AcademicStandards.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Standard not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)
    

def update_standard_data(user,data,standard_id):
    try:
        # Check if the academic board exists
        academic_standard = AcademicStandards.objects.get(id=standard_id)
        
        update_data = {field: value for field, value in data.dict().items() if value!= None}
        if update_data:
            for field, value in update_data.items():
                if field == "medium_id":
                    medium = AcademicMediums.objects.get(id=value)
                    academic_standard.medium_name = medium
                else:
                    setattr(academic_standard, field, value)
     
        academic_standard.save()
        updated_standard = {
            "id": str(academic_standard.id),
            "standard":academic_standard.standard,
            "medium_id": str(academic_standard.medium_name.id),
            "medium_name": academic_standard.medium_name.medium_name,
            "board_id": str(academic_standard.medium_name.board_name.id),
            "board_name": str(academic_standard.medium_name.board_name.board_name),
            "status": academic_standard.status,
            "created_at": academic_standard.created_at,
            "updated_at": academic_standard.updated_at,
        }

        response_data = {
            "result": True,
            "data": updated_standard,
            "message": "Academic stnadard updated successfully."
        }

        return response_data
    
    except AcademicStandards.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Standard not found"
        }
        return JsonResponse(response_data, status=400)
    
    except AcademicMediums.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Medium not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)
    

#-----------------------------------------------------------------------------------------------------------#
#--------------------------------------------------SUBJECT--------------------------------------------------#
#-----------------------------------------------------------------------------------------------------------#


def get_academic_subject_list(request, filter_prompt):
    try:
        academic_subjects = AcademicSubjects.objects.filter(standard__medium_name__board_name__business_owner=request.user).order_by('-created_at')
        print(academic_subjects,'asdfaf')
        q_objects = Q()
        if filter_prompt.search:
            q_objects = (
                Q(subject_name__icontains=filter_prompt.search) |
                Q(status__icontains=filter_prompt.search)
            )
            academic_subjects = academic_subjects.filter(q_objects)
            
        elif filter_prompt.status:
            q_objects &= Q(status=filter_prompt.status)
                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                     
        elif filter_prompt.medium_id and filter_prompt.board_id and filter_prompt.standard_id:
            q_objects &= (
                Q(standard__medium_name__id=filter_prompt.medium_id) &
                Q(standard__medium_name__board_name__id=filter_prompt.board_id) &
                Q(standard__id=filter_prompt.standard_id)
            )
        elif filter_prompt.medium_id and filter_prompt.board_id:
            q_objects &= (
                Q(standard__medium_name__id=filter_prompt.medium_id) &
                Q(standard__medium_name__board_name__id=filter_prompt.board_id)
            )
        elif filter_prompt.medium_id:
            q_objects &= Q(standard__medium_name__id=filter_prompt.medium_id)
        elif filter_prompt.board_id:
            q_objects &= Q(standard__medium_name__board_name__id=filter_prompt.board_id)
        elif filter_prompt.standard_id:
            q_objects &= Q(standard__id=filter_prompt.standard_id)

        elif filter_prompt.subject_id:
            q_objects &= Q(id=filter_prompt.subject_id)

        academic_subjects = academic_subjects.filter(q_objects)

        academic_subject_list = [
            {
                "id": str(subject.id),
                "subject_name": subject.subject_name,
                "board_id": str(subject.standard.medium_name.board_name.id),
                "board_name": subject.standard.medium_name.board_name.board_name,
                "medium_id": str(subject.standard.medium_name.id),
                "medium_name": subject.standard.medium_name.medium_name,
                "standard_id": str(subject.standard.id),
                "standard": subject.standard.standard,
                "status": subject.status,
                "created_at": subject.created_at,
                "updated_at": subject.updated_at,
            }
            for subject in academic_subjects
        ]
        paginated_subjects, items_per_page = paginate_data(request, academic_subject_list)

        return {
            "result": True,
            "data": list(paginated_subjects),
            "message": "Data retrieved successfully",   
            "pagination": {
                "page": paginated_subjects.number,
                "total_docs": paginated_subjects.paginator.count,
                "total_pages": paginated_subjects.paginator.num_pages,
                "per_page": items_per_page,
            },
        }
    

    except AcademicSubjects.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Subjects not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)
    

def add_subject_data(user, data):

    try:

        standard_id = data.standard_id 
    
        standard_instance = AcademicStandards.objects.get(id=standard_id) 
     # Convert to UUID and fetch the corresponding board instance
        subjects = AcademicSubjects.objects.create(
            subject_name=data.subject_name,
            standard=standard_instance,  # Use the fetched board instance
        )
    
        
        saved_subject = {
            "id": str(subjects.id),
            "subject_name":subjects.subject_name,
            "board_id": str(subjects.standard.medium_name.board_name.id),
            "board_name": subjects.standard.medium_name.board_name.board_name,
            "medium_id": str(subjects.standard.medium_name.id),
            "medium_name": subjects.standard.medium_name.medium_name,
            "standard_id":str(subjects.standard.id),
            "standard": subjects.standard.standard,
            "status": subjects.status,
            "created_at": subjects.created_at,
            "updated_at": subjects.updated_at,
        }
     
        response_data = {
            "result": True,
            "data": saved_subject,
            "message": "subject added successfully."
        }

        return response_data 
    
    except AcademicStandards.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Standard not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)


def get_academic_subject_data(subject_id):
    try:
        academic_subjects = AcademicSubjects.objects.get(id=subject_id)

        academic_standard_list = {
                "id": str(academic_subjects.id),
                "subject_name":academic_subjects.subject_name,
                "board_id": str(academic_subjects.standard.medium_name.board_name.id),
                "board_name": academic_subjects.standard.medium_name.board_name.board_name,
                "medium_id": str(academic_subjects.standard.medium_name.id),
                "medium_name": academic_subjects.standard.medium_name.medium_name,
                "standard_id":str(academic_subjects.standard.id),
                "standard": academic_subjects.standard.standard,
                "status": academic_subjects.status,
                "created_at": academic_subjects.created_at,
                "updated_at": academic_subjects.updated_at,
            }

        response_data = {
            "result": True,
            "data": academic_standard_list,
            "message": "Academic subject retrieved successfully."
        }

        return response_data
    
    except AcademicSubjects.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Subject not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)
    

def delete_subject_data(user,subject_id):
    try:
        # Check if the academic board exists
        academic_subject = AcademicSubjects.objects.get(id=subject_id)
        
        # Delete the academic board
        academic_subject.delete()

        response_data = {
            "result": True,
            "message": "Academic subject deleted successfully."
        }

        return response_data
    
    except AcademicSubjects.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Subject not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)
    

def update_subject_data(user,data,subject_id):
    try:
        # Check if the academic board exists
   
        academic_subjects = AcademicSubjects.objects.get(id=subject_id)

        update_data = {field: value for field, value in data.dict().items() if value!= None}
        if update_data:
            for field, value in update_data.items():
                if field == "standard_id":
                    standard = AcademicStandards.objects.get(id=value)
                    academic_subjects.standard = standard
                else:
                    setattr(academic_subjects, field, value)

        academic_subjects.save()
        updated_subject = {
            "id": str(academic_subjects.id),
            "subject_name":academic_subjects.subject_name,
            "board_id": str(academic_subjects.standard.medium_name.board_name.id),
            "board_name": academic_subjects.standard.medium_name.board_name.board_name,
            "medium_id": str(academic_subjects.standard.medium_name.id),
            "medium_name": academic_subjects.standard.medium_name.medium_name,
            "standard_id": str(academic_subjects.standard_id),
            "standard": academic_subjects.standard.standard,
            "status": academic_subjects.status,
            "created_at": academic_subjects.created_at,
            "updated_at": academic_subjects.updated_at,
        }

        response_data = {
            "result": True,
            "data": updated_subject,
            "message": "Academic subject updated successfully."
        }

        return response_data
    
    except AcademicSubjects.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Subject not found"
        }
        return JsonResponse(response_data, status=400)
    
    except AcademicStandards.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Standard not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)
    

#-----------------------------------------------------------------------------------------------------------#
#--------------------------------------------------CHAPTER--------------------------------------------------#
#-----------------------------------------------------------------------------------------------------------#


def get_academic_chapter_list(request, filter_prompt):
    try:
        academic_chapters = AcademicChapters.objects.filter(subject_name__standard__medium_name__board_name__business_owner=request.user).order_by('-created_at')
        

        if filter_prompt.search:
            q_objects = (
                Q(chapter_name__icontains=filter_prompt.search) |
                Q(subject_name__subject_name__icontains=filter_prompt.search) |
                Q(status__icontains=filter_prompt.search)
            )
            academic_chapters = academic_chapters.filter(q_objects)
        elif filter_prompt.board_id and filter_prompt.medium_id and filter_prompt.standard_id and filter_prompt.subject_id:
            academic_chapters = academic_chapters.filter(subject_name__id=filter_prompt.subject_id)

        elif filter_prompt.board_id and filter_prompt.medium_id and filter_prompt.standard_id:
            academic_chapters = academic_chapters.filter(subject_name__standard_id__id=filter_prompt.standard_id)

        elif filter_prompt.board_id and filter_prompt.medium_id:
            academic_chapters = academic_chapters.filter(subject_name__standard__medium_name__id=filter_prompt.medium_id)

        elif filter_prompt.status:
            academic_chapters = academic_chapters.filter(status=filter_prompt.status)

        elif filter_prompt.medium_id:
            academic_chapters = academic_chapters.filter(subject_name__standard__medium_name__id=filter_prompt.medium_id)

        elif filter_prompt.board_id:
            academic_chapters = academic_chapters.filter(subject_name__standard__medium_name__board_name__id=filter_prompt.board_id)

        elif filter_prompt.subject_id:
            academic_chapters = academic_chapters.filter(subject_name__id=filter_prompt.subject_id)

        elif filter_prompt.chapter_id:
            academic_chapters = academic_chapters.filter(id=filter_prompt.chapter_id)
        
        academic_chapters_list = []

        if filter_prompt.subject_ids:
            subject_ids_str = filter_prompt.subject_ids.strip()
            subject_ids = subject_ids_str.split(",")
            
            for subject_id in subject_ids:
                chapters_for_subject = academic_chapters.filter(subject_name__id=subject_id)
                subject_info = {
                    "subject_id": subject_id,
                    "subject_name": chapters_for_subject.first().subject_name.subject_name,
                    "chapters": [
                        {
                            "id": str(chapter.id),
                            "chapter_name": chapter.chapter_name,
                        }
                        for chapter in chapters_for_subject
                    ]
                }
                academic_chapters_list.append(subject_info)
            paginated_chapters, items_per_page = paginate_data(request, academic_chapters_list)
            return {
                "result": True,
                "data": list(paginated_chapters),
                "message": "Data retrieved successfully",   
                "pagination": {
                    "page": paginated_chapters.number,
                    "total_docs": paginated_chapters.paginator.count,
                    "total_pages": paginated_chapters.paginator.num_pages,
                    "per_page": items_per_page,
                },
            }
        else:
            academic_chapters_list = [
                {
                    "id": str(chapter.id),
                    "chapter_name": chapter.chapter_name,
                    "board_id": str(chapter.subject_name.standard.medium_name.board_name.id),
                    "board_name": chapter.subject_name.standard.medium_name.board_name.board_name,
                    "medium_id": str(chapter.subject_name.standard.medium_name.id),
                    "medium_name": chapter.subject_name.standard.medium_name.medium_name,
                    "standard_id": str(chapter.subject_name.standard_id),
                    "standard": chapter.subject_name.standard.standard,
                    "subject_id": str(chapter.subject_name.id),
                    "subject_name": chapter.subject_name.subject_name,
                    "status": chapter.status,
                    "created_at": chapter.created_at,
                    "updated_at": chapter.updated_at,
                }
                for chapter in academic_chapters
            ]
            paginated_chapters, items_per_page = paginate_data(request, academic_chapters_list)

            return {
                "result": True,
                "data": list(paginated_chapters),
                "message": "Data retrieved successfully",   
                "pagination": {
                    "page": paginated_chapters.number,
                    "total_docs": paginated_chapters.paginator.count,
                    "total_pages": paginated_chapters.paginator.num_pages,
                    "per_page": items_per_page,
                },
            }

      
    
    except AcademicChapters.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Chapter not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)
    

def add_chapter_data(user, data):

    try:

        subject_id = data.subject_id 
    
        subject_instance = AcademicSubjects.objects.get(id=subject_id) 
        chapter = AcademicChapters.objects.create(
            chapter_name=data.chapter_name,
            subject_name=subject_instance, 
        )
    
        saved_chapter = {
            "id": str(chapter.id),
            "chapter_name":chapter.chapter_name,
            "board_id": str(chapter.subject_name.standard.medium_name.board_name.id),
            "board_name": chapter.subject_name.standard.medium_name.board_name.board_name,
            "medium_id": str(chapter.subject_name.standard.medium_name.id),
            "medium_name": chapter.subject_name.standard.medium_name.medium_name,
            "standard_id": str(chapter.subject_name.standard_id),
            "standard": chapter.subject_name.standard.standard,
            "subject_id": str(chapter.subject_name.id),
            "subject_name": chapter.subject_name.subject_name,
            "status": chapter.status,
            "created_at": chapter.created_at,
            "updated_at": chapter.updated_at,
        }
     
        response_data = {
            "result": True,
            "data": saved_chapter,
            "message": "chapter added successfully."
        }

        return response_data 

    except AcademicSubjects.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Subject not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)
    

def get_academic_chapter_data(chapter_id):
    try:
        academic_chapters = AcademicChapters.objects.get(id=chapter_id)

        academic_chapter_list = {
                "id": str(academic_chapters.id),
                "chapter_name":academic_chapters.chapter_name,
                "board_id": str(academic_chapters.subject_name.standard.medium_name.board_name.id),
                "board_name": academic_chapters.subject_name.standard.medium_name.board_name.board_name,
                "medium_id": str(academic_chapters.subject_name.standard.medium_name.id),
                "medium_name": academic_chapters.subject_name.standard.medium_name.medium_name,
                "standard_id": str(academic_chapters.subject_name.standard_id),
                "standard": academic_chapters.subject_name.standard.standard,
                "subject_id": str(academic_chapters.subject_name.id),
                "subject_name": academic_chapters.subject_name.subject_name,
                "status": academic_chapters.status,
                "created_at": academic_chapters.created_at,
                "updated_at": academic_chapters.updated_at,
            }

        response_data = {
            "result": True,
            "data": academic_chapter_list,
            "message": "Academic subject retrieved successfully."
        }

        return response_data

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)
    


def delete_chapter_data(user,chapter_id):
    try:
        # Check if the academic board exists
        academic_chapter = AcademicChapters.objects.get(id=chapter_id)
        
        # Delete the academic board
        academic_chapter.delete()

        response_data = {
            "result": True,
            "message": "Academic chapter deleted successfully."
        }

        return response_data
    
    except AcademicChapters.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Chapter not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)
    


def update_chapter_data(user,data,chapter_id):
    try:
        # Check if the academic board exists
        academic_chapters = AcademicChapters.objects.get(id=chapter_id)
        
        update_data = {field: value for field, value in data.dict().items() if value!= None}
        if update_data:
            for field, value in update_data.items():
                if field == "subject_id":
                    subject = AcademicSubjects.objects.get(id=value)
                    academic_chapters.subject_name = subject
                else:
                    setattr(academic_chapters, field, value)

        # academic_medium.board_name = academic_board
       
        academic_chapters.save()
      
        updated_chapter = {
            "id": str(academic_chapters.id),
            "chapter_name":academic_chapters.chapter_name,
            "board_id": str(academic_chapters.subject_name.standard.medium_name.board_name.id),
            "board_name": academic_chapters.subject_name.standard.medium_name.board_name.board_name,
            "medium_id": str(academic_chapters.subject_name.standard.medium_name.id),
            "medium_name": academic_chapters.subject_name.standard.medium_name.medium_name,
            "standard_id": str(academic_chapters.subject_name.standard_id),
            "standard": academic_chapters.subject_name.standard.standard,
            "subject_id": str(academic_chapters.subject_name.id),
            "subject_name": academic_chapters.subject_name.subject_name,
            "status": academic_chapters.status,
            "created_at": academic_chapters.created_at,
            "updated_at": academic_chapters.updated_at,
        }

        response_data = {
            "result": True,
            "data": updated_chapter,
            "message": "Academic chapter updated successfully."
        }

        return response_data
    
    except AcademicChapters.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Chapter not found"
        }
        return JsonResponse(response_data, status=400)
    
    except AcademicSubjects.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Subject not found"
        }
        return JsonResponse(response_data, status=400)
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)
    

#-----------------------------------------------------------------------------------------------------------#
#-------------------------------------------------QUESTION--------------------------------------------------#
#-----------------------------------------------------------------------------------------------------------#


def add_question_data(user,data):
    try:
        options_data = data.options 
        options_instance = Options.objects.create(
            option1=options_data.option1,
            option2=options_data.option2,
            option3=options_data.option3,
            option4=options_data.option4
        )

        academic_chapter = AcademicChapters.objects.get(id=data.chapter_id)
        if data.question_image:
            # Handle the image data here
            image_data = base64.b64decode(data.question_image)
            timestamp = int(time.time())
            unique_filename = f"question_{timestamp}.png"
            academic_question_image = ContentFile(image_data, name=unique_filename)
        else:
            academic_question_image = None
        question = AcademicQuestions.objects.create(
            academic_chapter=academic_chapter,
            question=data.question,
            academic_question_image = academic_question_image,
            options=options_instance, 
            answer=data.answer,
            question_category=data.question_category,
            marks=data.marks,
            time_duration=data.time,
            business_owner=user
        )

        question_data = {
            "id": str(question.id),
            "question": question.question,
            "question_image":question.academic_question_image.url if question.academic_question_image else None,
            "answer": question.answer,
            "options": options_data,  # Return the options data as-is
            "board_id": str(question.academic_chapter.subject_name.standard.medium_name.board_name.id),
            "board_name": str(question.academic_chapter.subject_name.standard.medium_name.board_name.board_name),
            "medium_id": str(question.academic_chapter.subject_name.standard.medium_name.id),
            "medium_name": str(question.academic_chapter.subject_name.standard.medium_name.medium_name),
            "standard_id": str(question.academic_chapter.subject_name.standard.id),
            "standard_name": str(question.academic_chapter.subject_name.standard.standard),
            "subject_id": str(question.academic_chapter.subject_name.id),
            "subject_name": question.academic_chapter.subject_name.subject_name, 
            "chapter_id": str(question.academic_chapter.id),
            "chapter_name": question.academic_chapter.chapter_name,
            "question_category": question.question_category,
            "marks": str(question.marks),
            "time": question.time_duration,
            "status": question.status,
            "created_at":question.created_at,
            "updated_at":question.updated_at
        }

        response_data = {
            "result": True,
            "data": question_data,
            "message":"Question added successfully"
        }
        return response_data

    except AcademicChapters.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Chapter not found"
        }
        return JsonResponse(response_data, status=400)
    except Options.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Options not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)
    


def get_academic_question_list(request, filter_prompt):
    try:

      # Your base queryset
        base_query = AcademicQuestions.objects.filter(business_owner=request.user).order_by('-created_at')

        # Apply filters based on filter_prompt
        q_objects = Q()
        if filter_prompt:
            if filter_prompt.search:
                q_objects |= (
                    Q(question__icontains=filter_prompt.search) |
                    Q(answer__icontains=filter_prompt.search) |
                    Q(academic_chapter__chapter_name__icontains=filter_prompt.search) |
                    Q(academic_chapter__subject_name__subject_name__icontains=filter_prompt.search) |
                    Q(question_category__icontains=filter_prompt.search) |
                    Q(marks__icontains=filter_prompt.search) |
                    Q(time_duration__icontains=filter_prompt.search) |
                    Q(status__icontains=filter_prompt.search)
                )
            elif filter_prompt.status:
                q_objects &= Q(status=filter_prompt.status)
            elif filter_prompt.chapter_id:
                q_objects &= Q(academic_chapter__id=filter_prompt.chapter_id)
            elif filter_prompt.subject_id:
                q_objects &= Q(academic_chapter__subject_name=filter_prompt.subject_id)
            elif filter_prompt.standard_id:
                q_objects &= Q(academic_chapter__subject_name__standard__id=filter_prompt.standard_id)
            elif filter_prompt.medium_id:
                q_objects &= Q(academic_chapter__subject_name__standard__medium_name__id=filter_prompt.medium_id)
            elif filter_prompt.board_id:
                q_objects &= Q(academic_chapter__subject_name__standard__medium_name__board_name__id=filter_prompt.board_id)

        # Apply filters to the base query
        questions = base_query.filter(q_objects)

        # Use values to fetch only necessary fields
        question_list = questions.values(
            'id', 'question', 'academic_question_image', 'answer', 'marks', 'time_duration', 'status', 'created_at', 'updated_at', 'options_id', 'question_category',
            board_id=F('academic_chapter__subject_name__standard__medium_name__board_name__id'),
            board_name=F('academic_chapter__subject_name__standard__medium_name__board_name__board_name'),
            medium_id=F('academic_chapter__subject_name__standard__medium_name__id'),
            medium_name=F('academic_chapter__subject_name__standard__medium_name__medium_name'),
            standard_id=F('academic_chapter__subject_name__standard__id'),
            standard_name=F('academic_chapter__subject_name__standard__standard'),
            subject_id=F('academic_chapter__subject_name__id'),
            subject_name=F('academic_chapter__subject_name__subject_name'),
            chapter_id=F('academic_chapter__id'),
            chapter_name=F('academic_chapter__chapter_name'),
            # question_category=F('question_category'),
        )

        options_data = Options.objects.filter(id__in=question_list.values_list('options_id', flat=True))
        options_dict = {option.id: option for option in options_data}

        # Add options to the question_list
        for question_data in question_list:
            option_id = question_data['options_id']
            options_data = options_dict.get(option_id)
            if options_data:
                question_data['options'] = {
                    'option1': options_data.option1,
                    'option2': options_data.option2,
                    'option3': options_data.option3,
                    'option4': options_data.option4,
                }
            else:
                question_data['options'] = {}

        paginated_question_list,items_per_page = paginate_data(request, question_list)
        print(paginated_question_list)
        return {
            "result": True,
            "data": list(paginated_question_list),
            "message": "Data retrieved successfully",   
            "pagination": {
                "page": paginated_question_list.number,
                "total_docs": paginated_question_list.paginator.count,
                "total_pages": paginated_question_list.paginator.num_pages,
                "per_page": int(items_per_page)
            },
        } 
    
    except AcademicChapters.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Chapter not found"
        }
        return JsonResponse(response_data, status=400)
    
    except AcademicQuestions.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Question not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)
    

def get_academic_question_data(question_id):
    try:
        question = AcademicQuestions.objects.get(id=question_id)
        options_data = Options.objects.get(id=question.options_id)
        options_dict = {
                "option1": options_data.option1 if options_data else None,
                "option2": options_data.option2 if options_data else None,
                "option3": options_data.option3 if options_data else None,
                "option4": options_data.option4 if options_data else None,
            }
        
        question_data = {
                "id": str(question.id),
                "question": question.question,
                "question_image":question.academic_question_image.url if question.academic_question_image else None,
                "answer": question.answer,
                "options": options_dict,  # Return the options data as-is
                "board_id": str(question.academic_chapter.subject_name.standard.medium_name.board_name.id),
                "board_name": str(question.academic_chapter.subject_name.standard.medium_name.board_name.board_name),
                "medium_id": str(question.academic_chapter.subject_name.standard.medium_name.id),
                "medium_name": str(question.academic_chapter.subject_name.standard.medium_name.medium_name),
                "standard_id": str(question.academic_chapter.subject_name.standard.id),
                "standard_name": str(question.academic_chapter.subject_name.standard.standard),
                "subject_id": str(question.academic_chapter.subject_name.id),
                "subject_name": question.academic_chapter.subject_name.subject_name, 
                "chapter_id": str(question.academic_chapter.id),
                "chapter_name": question.academic_chapter.chapter_name,
                "question_category": question.question_category,
                "marks": str(question.marks),
                "time": str(question.time_duration),
                "status": question.status,
                "created_at":question.created_at,
                "updated_at":question.updated_at
            }
         
            
        response_data = {
            "result": True,
            "data": question_data,
            "message":"Questions retrived successfully"
        }
        return response_data
    
    except AcademicQuestions.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Question not found"
        }
        return JsonResponse(response_data, status=400)
    
    except Options.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Options not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)
    


def delete_question_data(user, question_id):
    try:
        # Check if the academic question exists
        academic_question = AcademicQuestions.objects.get(id=question_id)
        
        # Get the options associated with the question
        options_data = Options.objects.filter(id=academic_question.options_id)
        
        # Delete the academic question
        academic_question.delete()
        
        # Delete the associated options
        options_data.delete()

        response_data = {
            "result": True,
            "message": "Academic question and associated options deleted successfully."
        }

        return response_data

    except AcademicQuestions.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Question not found"
        }
        return JsonResponse(response_data, status=400)
    
    except Options.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Options not found"
        }
        return JsonResponse(response_data, status=400)
    
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)


def update_question_data(data, question_id):
    try:
        question = AcademicQuestions.objects.get(id=question_id)
        try:
            options_data = Options.objects.get(id=question.options_id)
        except Options.DoesNotExist:
            options_data = None

        update_data = {field: value for field, value in data.dict().items() if value is not None}
        print(update_data,"sdfsfs")
        if update_data:
            if "options" in update_data and options_data:
                new_options = update_data.pop("options")
                for field, value in new_options.items():
                    if hasattr(options_data, field) and value is not None:
                        setattr(options_data, field, value)
                options_data.save()

            if "question_image" in update_data and update_data["question_image"]:
                # Handle the image data here
                image_data = base64.b64decode(update_data["question_image"])
                timestamp = int(time.time())
                unique_filename = f"question_{timestamp}.png"
                academic_question_image = ContentFile(image_data, name=unique_filename)
                question.academic_question_image = academic_question_image

            for field, value in update_data.items():
                if field == "chapter_id":
                    chapter = AcademicChapters.objects.get(id=value)
                    question.academic_chapter = chapter
                if field == "time":
                    question.time_duration = value

                else:
                    setattr(question, field, value)
            question.save()
            question_data = {
                "id": str(question.id),
                "question": question.question,
                "question_image":question.academic_question_image.url if question.academic_question_image else None,
                "answer": question.answer,
                "board_id": str(question.academic_chapter.subject_name.standard.medium_name.board_name.id),
                "board_name": str(question.academic_chapter.subject_name.standard.medium_name.board_name.board_name),
                "medium_id": str(question.academic_chapter.subject_name.standard.medium_name.id),
                "medium_name": str(question.academic_chapter.subject_name.standard.medium_name.medium_name),
                "standard_id": str(question.academic_chapter.subject_name.standard.id),
                "standard_name": str(question.academic_chapter.subject_name.standard.standard),
                "subject_id": str(question.academic_chapter.subject_name.id),
                "subject_name": question.academic_chapter.subject_name.subject_name, 
                "chapter_id": str(question.academic_chapter.id),
                "chapter_name": question.academic_chapter.chapter_name,
                "question_category": question.question_category,
                "marks": str(question.marks),
                "time": str(question.time_duration),
                "status": question.status,
                "created_at": question.created_at,
                "updated_at": question.updated_at
            }
            # Construct response options_dict based on updated or existing options
            if options_data:
                options_dict = {
                    "option1": options_data.option1,
                    "option2": options_data.option2,
                    "option3": options_data.option3,
                    "option4": options_data.option4,
                }
                question_data["options"] = options_dict
            else:
                options_dict = {
                    "option1": question.options_data.option1,
                    "option2": question.options_data.option2,
                    "option3": question.options_data.option3,
                    "option4": question.options_data.option4,
                }
                question_data["options"] = options_dict

            

            response_data = {
                "result": True,
                "data": question_data,
                "message": "Question updated successfully"
            }
            return response_data

    except AcademicQuestions.DoesNotExist:
        response_data = {
                    "result": False,
                    "message": "Question not found"
                }
        return JsonResponse(response_data, status=400)

    except Options.DoesNotExist:
        response_data = {
                    "result": False,
                    "message": "Option not found"
                }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)
 
from django.core import serializers
def create_academic_exam(user, data):
    try:
     
        standard_instance = AcademicStandards.objects.get(id=data.standard_id)
        total_weightage = data.total_questions
        selected_acad_questions_set1 = []
        selected_acad_questions_set2 = []        
        selected_acad_questions_set3 = [] 

        exam_data_calculated = []
        start_time = time.time()           
        def backtrack(selected_questions, remaining_time, remaining_marks, remaining_easy_questions, remaining_medium_questions, remaining_hard_questions, question_data_list):
            taken_time = time.time() - start_time
            if taken_time > 120:  # minutes in seconds
                raise TimeoutError("Questions not found.")
            if remaining_time < 0 and remaining_marks < 0:
                return False
            if remaining_easy_questions == 0 and remaining_medium_questions == 0 and remaining_hard_questions == 0:
                if remaining_time == 0 and remaining_marks == 0:
                    return True

            for _ in range(len(question_data_list)):
                selected_question = random.choice(question_data_list)
                if selected_question in question_data_list:
                    question_data_list.remove(selected_question)
                    
                if selected_question.question_category == "easy" and remaining_easy_questions > 0 and selected_question.time_duration <= remaining_time and selected_question.marks <= remaining_marks:
                    selected_questions.append(selected_question)
                    updated_remaining_time = remaining_time - selected_question.time_duration
                    updated_remaining_marks = remaining_marks - selected_question.marks
                    updated_remaining_easy_questions = remaining_easy_questions - 1
                    # print("new", updated_remaining_easy_questions)      
                    # print("ADSA", selected_questions)
                    # print("timeee", updated_remaining_time)
                    # print("markssss", updated_remaining_marks)
                    if backtrack(selected_questions, updated_remaining_time, updated_remaining_marks, updated_remaining_easy_questions, remaining_medium_questions, remaining_hard_questions, question_data_list):
                        return True
                    else:
                        selected_questions.pop()
                        # print("SSSS")
                        return False
                    
                elif selected_question.question_category == "medium" and remaining_medium_questions > 0 and selected_question.time_duration <= remaining_time and selected_question.marks <= remaining_marks:
                    selected_questions.append(selected_question)
                    updated_remaining_time = remaining_time - selected_question.time_duration
                    updated_remaining_marks = remaining_marks - selected_question.marks
                    updated_remaining_medium_questions = remaining_medium_questions - 1
                    # print("new", updated_remaining_medium_questions)      
                    # print("ADSA", selected_questions)
                    # print("timeee", updated_remaining_time)
                    # print("markssss", updated_remaining_marks)
                    if backtrack(selected_questions, updated_remaining_time, updated_remaining_marks, remaining_easy_questions, updated_remaining_medium_questions, remaining_hard_questions, question_data_list):
                        return True
                    else:
                        selected_questions.pop()
                        # print("SSSS")
                        return False
                    
                elif selected_question.question_category == "hard" and remaining_hard_questions > 0 and selected_question.time_duration <= remaining_time and selected_question.marks <= remaining_marks:
                    selected_questions.append(selected_question)
                    updated_remaining_time = remaining_time - selected_question.time_duration
                    updated_remaining_marks = remaining_marks - selected_question.marks
                    updated_remaining_hard_questions = remaining_hard_questions - 1
                    # print("new", updated_remaining_hard_questions)      
                    # print("ADSA", selected_questions)
                    # print("timeee", updated_remaining_time)
                    # print("markssss", updated_remaining_marks)
                    if backtrack(selected_questions, updated_remaining_time, updated_remaining_marks, remaining_easy_questions, remaining_medium_questions, updated_remaining_hard_questions, question_data_list):
                        return True
                    else:
                        selected_questions.pop()
                        # print("SSSS")
                        return False

            return False

        subject_data_list = []
        
        for subject_data in data.exam_data:
            subject_weightage = sum(
                qtype for qtype in [subject_data.easy_question, subject_data.medium_question, subject_data.hard_question]
            )
            subject_percentage = round(subject_weightage / total_weightage, 2)

            subject_time = int(data.time_duration * subject_percentage + 0.5)
            subject_marks = round(float(data.total_marks * subject_percentage))
            # print(subject_marks, "MMMMAAAARRRKKKKKK")
            print(subject_data.subject_id)
            print(subject_time)
            print(subject_marks)
            subject_instance = AcademicSubjects.objects.get(id=subject_data.subject_id)
            subject_data_list.append({
            "subject_id": subject_data.subject_id,
            "subject_time": subject_time,
            "subject_marks": subject_marks,
            # Include other relevant data if needed
        })

            # chapter_instance = AcademicChapters.objects.filter(id__in=subject_data.chapter)
            # chapters = list(chapter_instance)
            # chapter_ids = [f"{item.id}," for item in chapters]
            # chapters = " ".join(chapter_ids)
            # question_data = (academic_chapter__subject_name=subject_instance)
            question_data = AcademicQuestions.objects.filter(academic_chapter__id__in=subject_data.chapter)
            question_data_list = list(question_data)
            
            selected_questions_set1 = []
            backtrack_result_set1 = backtrack(selected_questions_set1, subject_time, subject_marks, subject_data.easy_question, subject_data.medium_question, subject_data.hard_question, question_data_list)
        
            selected_questions_set2 = []
            backtrack_result_set2 = backtrack(selected_questions_set2, subject_time, subject_marks, subject_data.easy_question, subject_data.medium_question, subject_data.hard_question, question_data_list)
            
            selected_questions_set3 = []
            backtrack_result_set3 = backtrack(selected_questions_set3, subject_time, subject_marks, subject_data.easy_question, subject_data.medium_question, subject_data.hard_question, question_data_list)
        

            while not backtrack_result_set1 and len(selected_questions_set1) != subject_weightage:
                question_data_list = list(question_data)  
                selected_questions_set1 = []  
                remaining_marks = subject_marks 
                
                remaining_time = subject_time 
                backtrack_result_set1 = backtrack(selected_questions_set1, remaining_time, remaining_marks, subject_data.easy_question, subject_data.medium_question, subject_data.hard_question, question_data_list)

            while not backtrack_result_set2 and len(selected_questions_set2) != subject_weightage:
                question_data_list = list(question_data)  
                selected_questions_set2 = []  
                remaining_marks = subject_marks 
                
                remaining_time = subject_time 
                backtrack_result_set2 = backtrack(selected_questions_set2, remaining_time, remaining_marks, subject_data.easy_question, subject_data.medium_question, subject_data.hard_question, question_data_list)

            while not backtrack_result_set3 and len(selected_questions_set3) != subject_weightage:
                question_data_list = list(question_data)  
                selected_questions_set3 = []  
                remaining_marks = subject_marks 
                
                remaining_time = subject_time
                backtrack_result_set3 = backtrack(selected_questions_set3, remaining_time, remaining_marks, subject_data.easy_question, subject_data.medium_question, subject_data.hard_question, question_data_list)


            if backtrack_result_set1:
    
                for question in selected_questions_set1:
                    options_data = Options.objects.get(id=question.options_id)
                    options_dict = {
                            "option1": options_data.option1 if options_data else None,
                            "option2": options_data.option2 if options_data else None,
                            "option3": options_data.option3 if options_data else None,
                            "option4": options_data.option4 if options_data else None,
                        }
                    acad_exam_instance = AcadExam(
                        
                        id = str(question.id),
                        question=question.question,
                        time=float(question.time_duration),
                        mark=question.marks,
                        question_category=question.question_category,
                        subject = str(question.academic_chapter),
                        options = options_dict,
                        answer = question.answer,
                        question_image=question.academic_question_image.url if question.academic_question_image else None,
                        
                    )
                    selected_acad_questions_set1.append(acad_exam_instance)

            if backtrack_result_set2:
               
                for question in selected_questions_set2:
                    options_data = Options.objects.get(id=question.options_id)
                    options_dict = {
                            "option1": options_data.option1 if options_data else None,
                            "option2": options_data.option2 if options_data else None,
                            "option3": options_data.option3 if options_data else None,
                            "option4": options_data.option4 if options_data else None,
                        }
                    acad_exam_instance = AcadExam(
                        id = str(question.id),
                        question=question.question,
                        time=float(question.time_duration),
                        mark=question.marks,
                        question_category=question.question_category,
                        subject = str(question.academic_chapter),
                        options = options_dict,
                        answer = question.answer
                    )
                    selected_acad_questions_set2.append(acad_exam_instance)

            if backtrack_result_set3:
                for question in selected_questions_set3:
                    options_data = Options.objects.get(id=question.options_id)
                    options_dict = {
                            "option1": options_data.option1 if options_data else None,
                            "option2": options_data.option2 if options_data else None,
                            "option3": options_data.option3 if options_data else None,
                            "option4": options_data.option4 if options_data else None,
                        }
                    acad_exam_instance = AcadExam(
                        id = str(question.id),
                        question=question.question,
                        time=float(question.time_duration),
                        mark=question.marks,
                        question_category=question.question_category,
                        subject = str(question.academic_chapter),
                        options = options_dict,
                        answer = question.answer
                    )
                    selected_acad_questions_set3.append(acad_exam_instance)

            # print("------------------------------------------------------")

        
        result = {
            "subject_data": subject_data_list,
            "set1": selected_acad_questions_set1,
            "set2": selected_acad_questions_set2 if selected_acad_questions_set2 else None,
            "set3": selected_acad_questions_set3 if selected_acad_questions_set3 else None,
        }
      
        return result

    except TimeoutError as timeout_error:
        response_data = {
            "result": False,
            "message": "Backtracking took too long to complete"
        }
        return response_data
    
    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)

def get_acad_examlist(request, query):
    try: 
        exams = AcademicExams.objects.filter(business_owner=request.user, start_date__isnull=True).order_by('-created_at')
       
        if query.standard:
            exams = exams.filter(standard=query.standard)
        if query.subject:
            exams = exams.filter(exam_data__subject=query.subject)
        
        if query.chapter:
            exams = exams.filter(exam_data__chapter__contains=query.chapter)
        
        if query.search:
            search_terms = query.search.strip().split()
            search_query = Q()

            for term in search_terms:
                search_query |= (
                     Q(exam_title__icontains=term)
                    | Q(status__icontains=term)
                )
            exams = exams.filter(search_query)
       
        exam_list = []
        for exam in exams:
            exam_data_list = []
            for exam_data in exam.exam_data.all():
                subject_name = exam_data.subject.subject_name
                chapter = exam_data.chapter
                
                exam_data_list.append({"subject": subject_name, "chapters": chapter})

            exam_detail = {
                "id": str(exam.id),
                "exam_title": exam.exam_title,
                "board_id": str(exam.standard.medium_name.board_name.id),
                "board_name": exam.standard.medium_name.board_name.board_name,
                "medium_id": str(exam.standard.medium_name.id),
                "medium_name": exam.standard.medium_name.medium_name,
                "standard_id": str(exam.standard.id),
                "standard_name": exam.standard.standard,
                "total_question": exam.total_questions,
                "time_duration": exam.time_duration,
                "negative_marks": exam.negative_marks,
                "total_marks": exam.total_marks,
                "start_date": exam.start_date,
                "exam_data": exam_data_list, 
            }
            exam_list.append(exam_detail)
            
        paginated_examlist_data, items_per_page = paginate_data(request, exam_list)

        return {
            "result": True,
            "data": list(paginated_examlist_data),
            "message": "Data retrieved successfully",   
            "pagination": {
                "page": paginated_examlist_data.number,
                "total_docs": paginated_examlist_data.paginator.count,
                "total_pages": paginated_examlist_data.paginator.num_pages,
                "per_page": items_per_page,
            },
        }
    
    except AcademicExams.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Exam not found"
        }
        return JsonResponse(response_data, status=400)
    
    except Exception as e:
        response_data = {
            "result": False,
            "message": "Something went wrong"
        }
        return JsonResponse(response_data, status=400)
    


def start_acad_exam(data):
    try:
        exam_id = data.exam_id  # Assuming 'exam_id' is in the payload
        print(exam_id)
        if not exam_id:
            response_data = {
                "result": False,
                "message": "exam_id is required in the payload"
            }
            return JsonResponse(response_data, status=400)

        exam = AcademicExams.objects.get(id=exam_id)
        print(exam,"thisgigfg")
        exam.start_date = datetime.now()
        exam.save()
        result = {
            "result": True,
            "message": "Exam started."
        }

        return result

    except AcademicExams.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Exam not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
            "result": False,
            "message": print("Something went wrong")
        }
        return JsonResponse(response_data, status=400)



def get_acad_examreport(user, query):
    try: 
        
        exams = AcademicExams.objects.filter(business_owner=user)
        
        if query.standard:
            exams = exams.filter(standard=query.standard)
        if query.subject:
            exams = exams.filter(exam_data__subject=query.subject)
        
        if query.chapter:
            exams = exams.filter(exam_data__chapter__contains=query.chapter)
        
        if query.search:
            search_terms = query.search.strip().split()
            search_query = Q()

            for term in search_terms:
                search_query |= (
                     Q(exam_title__icontains=term)
                    | Q(status__icontains=term)
                    
                )
       
        exam_list = []
        for exam in exams:
            exam_data = {
                "id":str(exam.id),
                "exam_title": exam.exam_title,
                "standard": str(exam.standard.standard),
                "medium":exam.standard.medium_name.medium_name,
                "board":exam.standard.medium_name.board_name.board_name,
                "total_question":exam.total_questions,
                "time_duration":exam.time_duration,
                "negative_marks":exam.negative_marks,
                "total_marks":exam.total_marks,
               
            }
            exam_list.append(exam_data)
        return exam_list
    
    except AcademicExams.DoesNotExist:
        response_data = {
            "result": False,
            "message": "Exam not found"
        }
        return JsonResponse(response_data, status=400)

    except Exception as e:
        response_data = {
                    "result": False,
                    "message": "Something went wrong"
                }
        return JsonResponse(response_data, status=400)


def start_acad_CSExam(user, data):
    
    try:
        exam_data_calculated = []
        standard_id = data.standard_id  # Assuming 'standard_id' is a valid UUID4
        standard_instance = AcademicStandards.objects.get(id=standard_id)
        for subject_data in data.exam_data:
            easy=subject_data.easy_question
            medium=subject_data.medium_question
            hard=subject_data.hard_question
            subject_id = subject_data.subject_id
            subject_instance = AcademicSubjects.objects.get(id=subject_id)
            chapter_instance = AcademicChapters.objects.filter(id__in=subject_data.chapter)
            chapters = list(chapter_instance)
        # Create a new exam object with the provided data
        exam = AcademicExams.objects.create(
            exam_title=data.exam_title,
            standard=standard_instance,
            total_questions=data.total_questions,
            time_duration=data.time_duration,
            passing_marks=data.passing_marks,
            total_marks=data.total_marks,
            negative_marks=data.negative_marks,
            option_e=data.option_e,
            business_owner=user  
        )

        # Get a list of selected question IDs from the provided data
        selected_question_ids = data.question

        # Filter and get the actual question instances
        selected_questions = AcademicQuestions.objects.filter(id__in=selected_question_ids)
        # print(selected_questions)
        # print(data.time)
        # print(data.mark)
        # Add the selected questions to the exam
        for question in selected_questions:
            exam.question_set.add(question)
        # Save the exam
    
        # print(f"Exam {exam.id} saved")

        subject_data = data.subject_data
        # print(subject_data)
        for subject_info in subject_data:
            subject_id = subject_info.subject_id
            subject_time = subject_info.subject_time
            subject_marks = subject_info.subject_marks
            subject_instance = AcademicSubjects.objects.get(id=subject_id)
            # ... (your existing code)

            exam_data_instance = AcademicExamData(
                subject=subject_instance,
                easy_question=easy,
                chapter=chapters,
                medium_question=medium,
                hard_question=hard,
                time_per_subject=subject_time,
                marks_per_subject=subject_marks,
            )
            exam_data_instance.save()
        exam_data_calculated.append(exam_data_instance)
        exam.save()
        for exam_data_instance in exam_data_calculated:
            exam.exam_data.add(exam_data_instance)  
  
        # for exam in exam_data_calculated:
        #     exam_instance.exam_data.add(exam) 
        result = {
            "result": True,
            "message": "Exam created and will start soon",
            "exam_id": exam.id  # Include the exam_id in the response
        }

        return JsonResponse(result) 
    
    except Exception as e:
        response_data = {
            "result": False,
            "message": "Something went wrong"
        }
        return JsonResponse(response_data, status=400)


def get_results(exam_id):
    try:
        results = Results.objects.filter(Q(competitive_exam=exam_id) | Q(academic_exam=exam_id))
        return list(results)
    except Results.DoesNotExist:
        return None
    
def get_student(student_id):
    try:
        return Students.objects.get(id=student_id)
    except Students.DoesNotExist:
        return None
    
def get_competitive_exam(exam_id):
    try:
        return CompetitiveExams.objects.get(id=exam_id)
    except CompetitiveExams.DoesNotExist:
        return None
    except Exception as e:
        print(e)

def get_academic_exam(exam_id):
    try:
        return AcademicExams.objects.get(id=exam_id)
    except AcademicExams.DoesNotExist:
        return None
    
def get_exam_result(user,exam_id):
    try:
        print(user)
        # Get user business type from user
        business_type = user.business_type

        # Get exam details based on business type
        if business_type == "competitive":
            exam = get_competitive_exam(exam_id)
        elif business_type == 'academic':
            exam = get_academic_exam(exam_id)

        if not exam:
            return {"error": "Failed to fetch exam details."}

        exam_data = {
            "title": exam.exam_title,
            "passing_marks": exam.passing_marks,
            "total_marks": exam.total_marks,
            "time_duration": exam.time_duration,
            "total_questions": exam.total_questions
        }

        # Get exam results
        results_data = []

        for result in get_results(exam_id):
            student = get_student(result.student_id)
            if student:
                result_data = {
                    "rank": None,
                    "first_name": student.first_name or "N/A",
                    "last_name": student.last_name or "N/A",
                    "score": result.score,
                    "result": result.result,
                    "profile_image": student.profile_image.url if student.profile_image else None
                }
                results_data.append(result_data)

        sorted_results_data = sorted(results_data, key=lambda x: x['score'], reverse=True)
        for i, result in enumerate(sorted_results_data, start=1):
            result['rank'] = i

        result_data = {
            "exam_data": exam_data,
            "exam_results": sorted_results_data
        }

        return result_data

    except Exception as e:
        print(f"Error fetching results: {e}")
        return {"error": "An error occurred while fetching results."}


def get_presigned_key(filename):
   key = int(datetime.timestamp(datetime.now()))
   file_name = filename+'_'+str(key)
   return f"media/options_images/{filename}"

def create_presignedurl(file_name):
    try:
        file_name = file_name    
        ext = pathlib.Path(file_name).suffix
        prefix = pathlib.Path(file_name).stem
        key = get_presigned_key(prefix) + str(ext)
        try:
            response = s3_client.generate_presigned_post(
            Bucket=os.getenv('AWS_STORAGE_BUCKET_NAME'), Key=key, ExpiresIn=500
            )
            return {"status":200,"result": True, "data": response}
        except Exception as e:
            print(e)

    except Exception as e:
        return {"status":400,"result": False, "error": str(e)}
